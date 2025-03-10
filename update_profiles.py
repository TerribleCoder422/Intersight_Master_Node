#!/usr/bin/env python3
"""
Update Excel file for Intersight Foundation to support 8 profiles
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

# Path to the Excel file
file_path = '/Users/isaiahlaughlin/Lumos Dropbox/isaiah laughlin/Mac/Desktop/Python journey/Automate-Intersight/output/Intersight_Foundation.xlsx'

# Load the Excel file
workbook = load_workbook(file_path)

# Fix Profiles sheet to have 8 profiles
if 'Profiles' in workbook.sheetnames:
    # Store the index of the Profiles sheet to maintain order
    profiles_sheet_index = workbook.sheetnames.index('Profiles')
    profiles_sheet = workbook['Profiles']
    
    # Get existing headers and formatting
    headers_row = list(profiles_sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    
    # Get existing first profile for reference
    first_profile = list(profiles_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    
    # Add more profiles (keep first existing one, add 7 more)
    for i in range(2, 9):
        new_profile = list(first_profile)
        new_profile[0] = f"AI-Server-{i:02d}"  # Update profile name
        
        # Add row if it doesn't exist or update existing row
        if i <= profiles_sheet.max_row:
            for col, value in enumerate(new_profile, start=1):
                profiles_sheet.cell(row=i, column=col, value=value)
        else:
            profiles_sheet.append(new_profile)
    
    # Add organization dropdown validation to column C (as per memory requirements)
    organizations = ['default', 'Gruve']
    org_validation = DataValidation(
        type='list',
        formula1=f'"{",".join(organizations)}"',
        allow_blank=True
    )
    org_validation.add('C2:C9')  # Apply to Organization column (C) for 8 profiles
    profiles_sheet.add_data_validation(org_validation)
    
    # Add deploy dropdown validation
    deploy_validation = DataValidation(
        type='list',
        formula1='"Yes,No"',
        allow_blank=True
    )
    deploy_validation.add('G2:G9')  # Apply to Deploy column for 8 profiles
    profiles_sheet.add_data_validation(deploy_validation)
    
    # Add server dropdown with name and serial in column E (as per memory requirements)
    server_options = ['Server-01 (UCSX-210C-M6) | SN: FLM123456',
                     'Server-02 (UCSX-210C-M6) | SN: FLM123457',
                     'Server-03 (UCSX-210C-M6) | SN: FLM123458']
    server_validation = DataValidation(
        type='list',
        formula1=f'"{",".join(server_options)}"',
        allow_blank=True
    )
    server_validation.add('E2:E9')  # Apply to Server column (E) for 8 profiles
    profiles_sheet.add_data_validation(server_validation)
    
    print("Updated Profiles sheet with 8 profiles and proper dropdowns")

# For Policies sheet, ensure organization dropdown is in column D (as per memory requirements)
if 'Policies' in workbook.sheetnames:
    policies_sheet = workbook['Policies']
    organizations = ['default', 'Gruve']
    org_validation = DataValidation(
        type='list',
        formula1=f'"{",".join(organizations)}"',
        allow_blank=True
    )
    org_validation.add('D2:D50')  # Apply to Organization column (D)
    policies_sheet.add_data_validation(org_validation)
    print("Updated Policies sheet with organization dropdown in column D")

# For Template sheet, ensure organization dropdown is in column B (as per memory requirements)
if 'Template' in workbook.sheetnames:
    template_sheet = workbook['Template']
    organizations = ['default', 'Gruve']
    org_validation = DataValidation(
        type='list',
        formula1=f'"{",".join(organizations)}"',
        allow_blank=True
    )
    org_validation.add('B2:B50')  # Apply to Organization column (B)
    template_sheet.add_data_validation(org_validation)
    print("Updated Template sheet with organization dropdown in column B")

# No organization dropdown needed for Pools sheet (as per memory requirements)
print("No organization dropdown needed for Pools sheet")

# Save the workbook
workbook.save(file_path)
print(f"Saved updated Excel file to {file_path}")
