#!/usr/bin/env python3
"""
Fix Excel template to ensure the 'Policies' sheet contains the correct data 
and remove duplicate sheets.
"""

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

def fix_excel_template(file_path):
    """Fix Intersight Excel template sheet organization and dropdowns."""
    print(f"Opening workbook: {file_path}")
    wb = load_workbook(file_path)
    
    # Check for both Policies and Policies1 sheets
    print(f"Current sheets: {wb.sheetnames}")
    
    # If both Policies and Policies1 exist, we need to fix them
    if 'Policies' in wb.sheetnames and 'Policies1' in wb.sheetnames:
        # Get the sheets
        policies_sheet = wb['Policies']
        policies1_sheet = wb['Policies1']
        
        # Check which one has data
        policies_rows = list(policies_sheet.rows)
        policies1_rows = list(policies1_sheet.rows)
        
        policies_row_count = len(policies_rows)
        policies1_row_count = len(policies1_rows)
        
        print(f"Policies sheet has {policies_row_count} rows")
        print(f"Policies1 sheet has {policies1_row_count} rows")
        
        # If Policies1 has data and Policies is empty/has less data, copy from Policies1 to Policies
        if policies1_row_count > policies_row_count:
            print("Copying data from Policies1 to Policies...")
            
            # Clear the Policies sheet
            for row in policies_sheet.iter_rows():
                for cell in row:
                    cell.value = None
            
            # Copy data from Policies1 to Policies
            for row_idx, row in enumerate(policies1_sheet.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    policies_sheet.cell(row=row_idx, column=col_idx).value = cell.value
                    if cell.has_style:
                        policies_sheet.cell(row=row_idx, column=col_idx).font = cell.font
                        policies_sheet.cell(row=row_idx, column=col_idx).fill = cell.fill
                        policies_sheet.cell(row=row_idx, column=col_idx).border = cell.border
                        policies_sheet.cell(row=row_idx, column=col_idx).alignment = cell.alignment
            
            # Remove Policies1 sheet
            wb.remove(wb['Policies1'])
    
    # Do the same for other duplicate sheets
    for sheet_pair in [('Template', 'Template1'), ('Profiles', 'Profiles1')]:
        orig_sheet, dup_sheet = sheet_pair
        if orig_sheet in wb.sheetnames and dup_sheet in wb.sheetnames:
            print(f"Handling {orig_sheet} and {dup_sheet}...")
            orig = wb[orig_sheet]
            dup = wb[dup_sheet]
            
            orig_rows = len(list(orig.rows))
            dup_rows = len(list(dup.rows))
            
            if dup_rows > orig_rows:
                print(f"Copying data from {dup_sheet} to {orig_sheet}...")
                # Clear the original sheet
                for row in orig.iter_rows():
                    for cell in row:
                        cell.value = None
                
                # Copy data from duplicate to original
                for row_idx, row in enumerate(dup.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        orig.cell(row=row_idx, column=col_idx).value = cell.value
                        if cell.has_style:
                            orig.cell(row=row_idx, column=col_idx).font = cell.font
                            orig.cell(row=row_idx, column=col_idx).fill = cell.fill
                            orig.cell(row=row_idx, column=col_idx).border = cell.border
                            orig.cell(row=row_idx, column=col_idx).alignment = cell.alignment
                
                # Remove duplicate sheet
                wb.remove(dup)
    
    # Ensure organization dropdowns are in the correct columns
    # Create sample lists for dropdowns
    org_list = ["default", "DevOps", "Production", "Test", "UAT"]
    server_list = ["Server-1 (FCH1234V5Z7)", "Server-2 (FCH5678A9BC)", "Server-3 (FCH9012D3EF)"]
    
    # Profiles sheet: Organization in column C, Server dropdown in column E
    if 'Profiles' in wb.sheetnames:
        profiles_sheet = wb['Profiles']
        
        # Add organization dropdown to column C
        org_validation_profiles = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        profiles_sheet.add_data_validation(org_validation_profiles)
        org_validation_profiles.add('C2:C1000')  # Column C
        
        # Add server dropdown to column E
        server_validation = DataValidation(type='list', formula1=f'"{",".join(server_list)}"', allow_blank=True)
        profiles_sheet.add_data_validation(server_validation)
        server_validation.add('E2:E1000')  # Column E
    
    # Policies sheet: Organization in column D
    if 'Policies' in wb.sheetnames:
        policies_sheet = wb['Policies']
        
        # Add organization dropdown to column D
        org_validation_policies = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        policies_sheet.add_data_validation(org_validation_policies)
        org_validation_policies.add('D2:D1000')  # Column D
    
    # Template sheet: Organization in column B
    if 'Template' in wb.sheetnames:
        template_sheet = wb['Template']
        
        # Add organization dropdown to column B
        org_validation_template = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        template_sheet.add_data_validation(org_validation_template)
        org_validation_template.add('B2:B1000')  # Column B
    
    # Save the workbook
    fixed_file_path = file_path.replace('.xlsx', '_fixed.xlsx')
    wb.save(fixed_file_path)
    print(f"Fixed workbook saved to: {fixed_file_path}")
    return fixed_file_path

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = 'intersight_simplified_template.xlsx'
    
    fix_excel_template(file_path)
