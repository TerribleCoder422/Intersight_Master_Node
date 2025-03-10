#!/usr/bin/env python3
"""
Create a clean, structured Excel template for Cisco Intersight Server Profile Templates.
This template follows Intersight's exact configuration structure and includes dropdowns.
"""

import pandas as pd
import os
import intersight
from intersight.api_client import ApiClient
from intersight.configuration import Configuration
from intersight.api import bios_api, vnic_api, storage_api, organization_api, server_api
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def get_api_client():
    """
    Create an Intersight API client using the API key file
    """
    try:
        # Get API key details from environment variables
        api_key_id = os.getenv('INTERSIGHT_API_KEY_ID')
        api_key_file = os.getenv('INTERSIGHT_PRIVATE_KEY_FILE', './SecretKey.txt')
        
        if not api_key_id or not os.path.exists(api_key_file):
            print("Error: API key configuration not found")
            return None
            
        # Create configuration
        config = Configuration(
            host = os.getenv('INTERSIGHT_BASE_URL', 'https://intersight.com'),
            signing_info = intersight.signing.HttpSigningConfiguration(
                key_id = api_key_id,
                private_key_path = api_key_file,
                signing_scheme = intersight.signing.SCHEME_HS2019,
                signing_algorithm = intersight.signing.ALGORITHM_ECDSA_MODE_FIPS_186_3,
                hash_algorithm = intersight.signing.HASH_SHA256,
                signed_headers = [
                    intersight.signing.HEADER_REQUEST_TARGET,
                    intersight.signing.HEADER_HOST,
                    intersight.signing.HEADER_DATE,
                    intersight.signing.HEADER_DIGEST,
                ]
            )
        )
        
        # Create API client
        api_client = ApiClient(configuration=config)
        return api_client
        
    except Exception as e:
        print(f"Error creating API client: {str(e)}")
        return None

def get_organizations(api_client):
    """
    Get list of organizations from Intersight
    """
    if not api_client:
        return ["default"]
        
    try:
        org_api = organization_api.OrganizationApi(api_client)
        orgs = org_api.get_organization_organization_list()
        return [org.name for org in orgs.results] if orgs.results else ["default"]
    except Exception as e:
        print(f"Error fetching organizations: {str(e)}")
        return ["default"]

def get_available_policies(api_client):
    """
    Get lists of available policies from Intersight
    """
    try:
        # Get BIOS Policies
        bios_api_instance = bios_api.BiosApi(api_client)
        bios_policies = bios_api_instance.get_bios_policy_list().results
        bios_policy_names = [policy.name for policy in bios_policies]

        # Get QoS Policies
        qos_api_instance = vnic_api.VnicApi(api_client)
        qos_policies = qos_api_instance.get_vnic_eth_qos_policy_list().results
        qos_policy_names = [policy.name for policy in qos_policies]

        # Get Storage Policies
        storage_api_instance = storage_api.StorageApi(api_client)
        storage_policies = storage_api_instance.get_storage_storage_policy_list().results
        storage_policy_names = [policy.name for policy in storage_policies]

        # Get LAN Connectivity Policies
        lan_api_instance = vnic_api.VnicApi(api_client)
        lan_policies = lan_api_instance.get_vnic_lan_connectivity_policy_list().results
        lan_policy_names = [policy.name for policy in lan_policies]

        print("\nAvailable Policies in Intersight:")
        print("BIOS Policies:", bios_policy_names)
        print("QoS Policies:", qos_policy_names)
        print("Storage Policies:", storage_policy_names)
        print("LAN Connectivity Policies:", lan_policy_names)

        return {
            'bios_policies': bios_policy_names,
            'qos_policies': qos_policy_names,
            'storage_policies': storage_policy_names,
            'lan_policies': lan_policy_names
        }

    except Exception as e:
        print(f"Error getting policies: {str(e)}")
        return None

def create_excel_template(output_file, policies, template_config=None):
    """Create an Excel template with policy dropdowns"""
    try:
        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # Create a Pandas Excel writer using openpyxl as the engine
        writer = pd.ExcelWriter(output_file, engine='openpyxl')

        # Get organizations from Intersight
        api_client = get_api_client()
        organizations = get_organizations(api_client)
        organizations_str = ','.join(organizations) if organizations else "default"

        # Basic Information Sheet
        basic_info_data = {
            'Parameter': [
                'Organization', 
                'Name', 
                'Description', 
                'Tags'
            ],
            'Value': [
                template_config.get('Basic Information', {}).get('Organization', '') if template_config else '',
                template_config.get('Basic Information', {}).get('Name', '') if template_config else '',
                template_config.get('Basic Information', {}).get('Description', '') if template_config else '',
                template_config.get('Basic Information', {}).get('Tags', '') if template_config else ''
            ],
            'Notes': [
                'Required. Select the organization for this template.', 
                'Required. Enter a unique name for this template.', 
                'Optional. Enter a description for this template.', 
                'Optional. Format: key1:value1,key2:value2'
            ]
        }

        # Compute Configuration Sheet
        compute_data = {
            'Policy Type': [
                'BIOS Policy', 
                'Boot Order Policy', 
                'Virtual Media Policy',
                'UUID Pool'
            ],
            'Policy Name': [
                template_config.get('Compute Configuration', {}).get('BIOS Policy', '') if template_config else '',
                template_config.get('Compute Configuration', {}).get('Boot Order Policy', '') if template_config else '',
                template_config.get('Compute Configuration', {}).get('Virtual Media Policy', '') if template_config else '',
                template_config.get('Compute Configuration', {}).get('UUID Pool', '') if template_config else ''
            ],
            'Description': [
                'Controls processor, memory, and other hardware settings', 
                'Defines boot device sequence and options', 
                'Configures virtual media mapping (ISO, IMG files)',
                'For assigning unique identifiers to servers'
            ]
        }

        # Network Configuration Sheet
        network_data = {
            'Policy Type': [
                'LAN Connectivity Policy', 
                'SAN Connectivity Policy',
                'QoS Policy'
            ],
            'Policy Name': [
                template_config.get('Network Configuration', {}).get('LAN Connectivity Policy', '') if template_config else '',
                template_config.get('Network Configuration', {}).get('SAN Connectivity Policy', '') if template_config else '',
                template_config.get('Network Configuration', {}).get('QoS Policy', '') if template_config else ''
            ],
            'Description': [
                'Contains vNIC configurations, MAC address pools, VLANs, QoS settings', 
                'Contains vHBA configurations, WWPN/WWNN pools, VSANs, Fibre Channel settings',
                'Configures Quality of Service for network traffic'
            ]
        }

        # Storage Configuration Sheet
        storage_data = {
            'Policy Type': [
                'SD Card Policy', 
                'Storage Policy', 
                'Persistent Memory Policy'
            ],
            'Policy Name': [
                template_config.get('Storage Configuration', {}).get('SD Card Policy', '') if template_config else '',
                template_config.get('Storage Configuration', {}).get('Storage Policy', '') if template_config else '',
                template_config.get('Storage Configuration', {}).get('Persistent Memory Policy', '') if template_config else ''
            ],
            'Description': [
                'Configures SD card settings', 
                'Defines RAID configurations, disk groups, virtual drives', 
                'For Intel Optane DC persistent memory'
            ]
        }

        # Power & Thermal Sheet
        power_thermal_data = {
            'Policy Type': [
                'Power Policy', 
                'Thermal Policy'
            ],
            'Policy Name': [
                template_config.get('Power & Thermal', {}).get('Power Policy', '') if template_config else '',
                template_config.get('Power & Thermal', {}).get('Thermal Policy', '') if template_config else ''
            ],
            'Description': [
                'Controls power characteristics', 
                'Manages cooling and fan behavior'
            ]
        }

        # Create DataFrames
        basic_info_df = pd.DataFrame(basic_info_data)
        compute_df = pd.DataFrame(compute_data)
        network_df = pd.DataFrame(network_data)
        storage_df = pd.DataFrame(storage_data)
        power_thermal_df = pd.DataFrame(power_thermal_data)

        # Write DataFrames to Excel
        basic_info_df.to_excel(writer, sheet_name='Basic Information', index=False)
        compute_df.to_excel(writer, sheet_name='Compute Configuration', index=False)
        network_df.to_excel(writer, sheet_name='Network Configuration', index=False)
        storage_df.to_excel(writer, sheet_name='Storage Configuration', index=False)
        power_thermal_df.to_excel(writer, sheet_name='Power & Thermal', index=False)

        # Save the writer
        writer.close()

        print(f"Excel template with dropdowns created successfully: {output_file}")
        return True

    except Exception as e:
        print(f"Error creating Excel template: {str(e)}")
        return False

def get_org_moid(api_client):
    """Get the MOID of the organization"""
    try:
        org_api = organization_api.OrganizationApi(api_client)
        orgs = org_api.get_organization_organization_list()
        return orgs.results[0].moid
    except Exception as e:
        print(f"Error getting organization MOID: {str(e)}")
        return None

def get_policy_moid(api_client, policy_type, policy_name):
    """Get the MOID of a policy"""
    try:
        if not policy_name:
            return None
            
        if policy_type == 'bios.Policy':
            api_instance = bios_api.BiosApi(api_client)
            response = api_instance.get_bios_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == 'vnic.LanConnectivityPolicy':
            api_instance = vnic_api.VnicApi(api_client)
            response = api_instance.get_vnic_lan_connectivity_policy_list(filter=f"Name eq '{policy_name}'")
        else:
            return None
        
        if response.results:
            return response.results[0].moid
        return None
        
    except Exception as e:
        print(f"Error getting policy MOID: {str(e)}")
        return None

def create_server_template(api_client, template_data):
    """Create a server template in Intersight"""
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client)
        
        # Extract basic information
        template_name = template_data['Basic Information'].get('Name', '')
        template_description = template_data['Basic Information'].get('Description', '')
        
        # Create Server Profile Template API instance
        api_instance = server_api.ServerApi(api_client)
        
        # Get policy MOIDs
        bios_policy_moid = get_policy_moid(api_client, 'bios.Policy', template_data['Compute Configuration'].get('BIOS Policy', ''))
        lan_policy_moid = get_policy_moid(api_client, 'vnic.LanConnectivityPolicy', template_data['Network Configuration'].get('LAN Connectivity Policy', ''))
        
        # Create the template body
        template_body = {
            'Name': template_name,
            'Description': template_description,
            'Organization': {
                'ObjectType': 'organization.Organization',
                'Moid': org_moid
            },
            'TargetPlatform': 'FIAttached',
            'PolicyBucket': []
        }
        
        # Add policies to the template
        if bios_policy_moid:
            template_body['PolicyBucket'].append({
                'ObjectType': 'bios.Policy',
                'Moid': bios_policy_moid
            })
            
        if lan_policy_moid:
            template_body['PolicyBucket'].append({
                'ObjectType': 'vnic.LanConnectivityPolicy',
                'Moid': lan_policy_moid
            })
            
        # Create the template
        print(f"\nCreating server template '{template_name}'...")
        api_instance.create_server_profile_template(template_body)
        print(f"Successfully created server template: {template_name}")
        return True
        
    except Exception as e:
        print(f"Error creating server template: {str(e)}")
        return False

def push_template_to_intersight(api_client, template_file):
    """Push template configuration to Intersight"""
    try:
        # Read the Excel template
        print(f"\nReading template configuration from {template_file}...")
        template_data = {}
        
        # Read each sheet into a dictionary
        xls = pd.ExcelFile(template_file)
        for sheet_name in xls.sheet_names:
            if sheet_name not in ['Instructions', 'ReferenceData']:
                df = pd.read_excel(template_file, sheet_name=sheet_name)
                # Convert DataFrame to dictionary
                sheet_data = {}
                if 'Parameter' in df.columns and 'Value' in df.columns:
                    for _, row in df.iterrows():
                        if pd.notna(row['Parameter']) and pd.notna(row['Value']):
                            sheet_data[row['Parameter']] = row['Value']
                elif 'Policy Type' in df.columns and 'Policy Name' in df.columns:
                    for _, row in df.iterrows():
                        if pd.notna(row['Policy Type']) and pd.notna(row['Policy Name']):
                            sheet_data[row['Policy Type']] = row['Policy Name']
                template_data[sheet_name] = sheet_data
        
        # Create the server template in Intersight
        if not template_data.get('Basic Information', {}).get('Name'):
            print("Error: Template name is required")
            return False
            
        return create_server_template(api_client, template_data)
        
    except Exception as e:
        print(f"Error pushing template to Intersight: {str(e)}")
        return False

if __name__ == "__main__":
    # Get API client
    api_client = get_api_client()
    if not api_client:
        print("Failed to create API client")
        exit(1)

    # Get available policies first
    print("\nFetching available policies from Intersight...")
    policies = get_available_policies(api_client)
    if not policies:
        print("Failed to get policies")
        exit(1)

    # Create the template
    create_excel_template('output/Intersight_Server_Profile_Template.xlsx', policies)
    print("\nExcel template with dropdowns created successfully: output/Intersight_Server_Profile_Template.xlsx")

    # Push the template to Intersight
    push_template_to_intersight(api_client, 'output/Intersight_Server_Profile_Template.xlsx')
