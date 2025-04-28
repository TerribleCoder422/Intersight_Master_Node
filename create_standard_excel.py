#!/usr/bin/env python3
"""
Create a standard Excel template with basic data validation.
Uses the simplest approach that will actually work in Excel.
"""

import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_standard_excel(excel_file):
    """Create a simple Excel template with standard dropdowns"""
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Create sheets in correct order
        for sheet_name in ['Pools', 'Policies', 'Template', 'Profiles', 'Templates', 'Organizations', 'Servers']:
            wb.create_sheet(sheet_name)
        
        # Define header style
        header_fill = PatternFill(start_color='A0D7BE', end_color='A0D7BE', fill_type='solid')
        header_font = Font(bold=True)
        
        # Sample data - generic until update script runs
        orgs = ["default", "Organization-2", "Organization-3", "Organization-4"]
        resource_groups = ["default", "Resource-Group-2", "Resource-Group-3", "Resource-Group-4"]
        server_options = [
            "WMP2528012M | C220M5-Hosting-Server1", 
            "WMP2528012G | C220M5-Hosting-Server2",
            "WMP25280129 | C220M5-Hosting-Server3",
            "FCH2342W02W | C480MLM5-RH-CP-Worker1"
        ]
        
        # Create the Profiles sheet
        profiles = wb["Profiles"]
        
        # Add headers
        headers = ["Profile Name*", "Description", "Organization*", "Resource Group*", 
                   "Template Name*", "Server*", "Notes", "Deploy*"]
        
        for col, header in enumerate(headers, 1):
            cell = profiles.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            
        # Set column widths
        profiles.column_dimensions['A'].width = 25  # Profile Name
        profiles.column_dimensions['B'].width = 20  # Description
        profiles.column_dimensions['C'].width = 15  # Organization
        profiles.column_dimensions['D'].width = 20  # Resource Group
        profiles.column_dimensions['E'].width = 25  # Template Name
        profiles.column_dimensions['F'].width = 40  # Server
        profiles.column_dimensions['G'].width = 30  # Notes
        profiles.column_dimensions['H'].width = 10  # Deploy
        
        # Create sample data rows
        sample_rows = [
            ["AI-Server-01", "", "default", "AI POD Servers", "AI_POD_Template", "", "Production AI POD Host 1", "No"],
            ["AI-Server-02", "", "default", "ML Servers", "AI_POD_Template", "", "Production AI POD Host 2", "No"],
            ["AI-Server-03", "", "default", "DevOps", "AI_POD_Template", "", "Production AI POD Host 3", "No"],
            ["AI-Server-04", "", "default", "Production", "AI_POD_Template", "", "Production AI POD Host 4", "No"]
        ]
        
        for row_idx, row_data in enumerate(sample_rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                profiles.cell(row=row_idx, column=col_idx, value=value)
        
        # Create simple dropdowns (no dynamic formulas, just plain lists)
        # Organization dropdown - use more compatible format with comma instead of semicolon
        org_validation = DataValidation(type='list', formula1=f'"{",".join(orgs)}"', allow_blank=True)
        org_validation.add('C2:C1000')
        profiles.add_data_validation(org_validation)
        
        # Resource Group dropdown
        rg_validation = DataValidation(type='list', formula1=f'"{",".join(resource_groups)}"', allow_blank=True)
        rg_validation.add('D2:D1000')
        profiles.add_data_validation(rg_validation)
        
        # Server dropdown - limit the size for better compatibility
        # Only include first few servers to prevent Excel validation issues
        visible_servers = server_options[:10] if len(server_options) > 10 else server_options
        server_validation = DataValidation(type='list', formula1=f'"{",".join(visible_servers)}"', allow_blank=True)
        server_validation.add('F2:F1000')
        profiles.add_data_validation(server_validation)
        
        # Deploy dropdown - simpler validation
        deploy_validation = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
        deploy_validation.add('H2:H1000')
        profiles.add_data_validation(deploy_validation)
        
        # Pools sheet
        pools = wb["Pools"]
        pool_headers = ["Pool Type*", "Pool Name*", "Description", "First ID", "Size"]
        
        # Set column widths for Pools sheet
        pools.column_dimensions['A'].width = 20  # Pool Type
        pools.column_dimensions['B'].width = 30  # Pool Name
        pools.column_dimensions['C'].width = 40  # Description
        pools.column_dimensions['D'].width = 20  # First ID
        pools.column_dimensions['E'].width = 15  # Size
        
        for col, header in enumerate(pool_headers, 1):
            cell = pools.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            
        # Pools dropdown
        pool_types = ["MAC Pool", "UUID Pool"]
        pools_validation = DataValidation(type='list', formula1=f'"{",".join(pool_types)}"', allow_blank=True)
        pools_validation.add('A2:A1000')
        pools.add_data_validation(pools_validation)
        
        # Sample pools data with valid addresses for immediate push capability
        sample_pools = [
            ("MAC Pool", "AI_POD-MAC-A", "MAC Pool for AI POD Fabric A", "00:25:B5:A0:00:00", "256"),
            ("MAC Pool", "AI_POD-MAC-B", "MAC Pool for AI POD Fabric B", "00:25:B5:B0:00:00", "256"),
            ("UUID Pool", "AI_POD-UUID-Pool", "UUID Pool for AI POD Servers", "0000-000000000001", "100")
        ]
        
        # Add explanatory headers to make editing more intuitive
        for col, header in enumerate(["Pool Type*", "Pool Name*", "Description", "First Address*", "Size*"], 1):
            pools.cell(row=1, column=col).value = header
        
        for row_idx, row_data in enumerate(sample_pools, 2):
            for col_idx, value in enumerate(row_data, 1):
                pools.cell(row=row_idx, column=col_idx, value=value)
        
        # Policies sheet
        policies = wb["Policies"]
        policies_headers = ["Policy Type*", "Policy Name*", "Description", "Organization*"]
        
        # Set column widths for Policies sheet
        policies.column_dimensions['A'].width = 20  # Policy Type
        policies.column_dimensions['B'].width = 30  # Policy Name
        policies.column_dimensions['C'].width = 40  # Description
        policies.column_dimensions['D'].width = 20  # Organization
        
        for col, header in enumerate(policies_headers, 1):
            cell = policies.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            
        # Policy type dropdown
        policy_types = [
            "BIOS Policy",
            "Boot Order Policy",
            "vNIC / LAN Connectivity Policy",
            "vHBA / SAN Connectivity Policy", 
            "Local Disk Policy",
            "Storage Policy",
            "IMC Access Policy",
            "Power and Thermal Policy",
            "GPU Policy",
            "vMedia Policy",
            "IPMI Policy",
            "KVM Policy",
            "Serial-over-LAN Policy",
            "QoS Policy"
        ]
        policy_validation = DataValidation(type='list', formula1=f'"{",".join(policy_types)}"', allow_blank=True)
        policy_validation.add('A2:A1000')
        policies.add_data_validation(policy_validation)
        
        # Organization dropdown for policies
        org_validation_policies = DataValidation(type='list', formula1=f'"{",".join(orgs)}"', allow_blank=True)
        org_validation_policies.add('D2:D1000')
        policies.add_data_validation(org_validation_policies)
        
        # Sample policies with updated policy types
        sample_policies = [
            ("BIOS Policy", "AI_POD-BIOS", "Optimizes CPU, NUMA, and memory for AI workloads", "default"),
            ("Boot Order Policy", "AI_POD-BOOT", "PXE boot with local disk fallback for AI nodes", "default"),
            ("vNIC / LAN Connectivity Policy", "AI_POD-vNIC-A", "vNIC Policy for AI POD Fabric A with VLAN mappings", "default"),
            ("vNIC / LAN Connectivity Policy", "AI_POD-vNIC-B", "vNIC Policy for AI POD Fabric B with VLAN mappings", "default"),
            ("vHBA / SAN Connectivity Policy", "AI_POD-SAN", "FC configuration for large dataset storage access", "default"),
            ("Local Disk Policy", "AI_POD-Local-Storage", "RAID configuration for local SSDs", "default"),
            ("Storage Policy", "AI_POD-NVMe", "NVMe storage optimization for high IOPS", "default"),
            ("IMC Access Policy", "AI_POD-IMC", "Secure out-of-band management via CIMC", "default"),
            ("Power and Thermal Policy", "AI_POD-Thermal", "High-performance cooling profile for GPUs", "default"),
            ("GPU Policy", "AI_POD-GPU", "NVIDIA GPU configuration with SR-IOV enabled", "default"),
            ("vMedia Policy", "AI_POD-vMedia", "ISO attachment for automated OS provisioning", "default"),
            ("QoS Policy", "AI_POD-QoS", "Network QoS optimization for AI traffic", "default")
        ]
        
        for row_idx, row_data in enumerate(sample_policies, 2):
            for col_idx, value in enumerate(row_data, 1):
                policies.cell(row=row_idx, column=col_idx, value=value)
        
        # Template sheet
        template = wb["Template"]
        template_headers = ["Template Name*", "Organization*", "Resource Group*", 
                           "Description", "Target Platform*"]
        
        # Set column widths for Template sheet
        template.column_dimensions['A'].width = 30  # Template Name
        template.column_dimensions['B'].width = 20  # Organization
        template.column_dimensions['C'].width = 25  # Resource Group
        template.column_dimensions['D'].width = 40  # Description
        template.column_dimensions['E'].width = 20  # Target Platform
        
        for col, header in enumerate(template_headers, 1):
            cell = template.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            
        # Template sample data
        template_data = ["Ai_POD_Template", "default", "AI POD Servers", 
                         "Template for AI POD Servers", "FIAttached"]
        
        for col, value in enumerate(template_data, 1):
            template.cell(row=2, column=col, value=value)
            
        # Template dropdowns
        # Organization for template
        org_validation_template = DataValidation(type='list', formula1=f'"{",".join(orgs)}"', allow_blank=True)
        org_validation_template.add('B2:B1000')
        template.add_data_validation(org_validation_template)
        
        # Resource Group for template
        rg_validation_template = DataValidation(type='list', formula1=f'"{",".join(resource_groups)}"', allow_blank=True)
        rg_validation_template.add('C2:C1000')
        template.add_data_validation(rg_validation_template)
        
        # Target Platform dropdown
        platforms = ["FIAttached", "Standalone"]
        platform_validation = DataValidation(type='list', formula1=f'"{",".join(platforms)}"', allow_blank=True)
        platform_validation.add('E2:E1000')
        template.add_data_validation(platform_validation)
        
        # Servers sheet
        servers = wb["Servers"]
        servers_headers = ["Server Name", "Serial Number"]
        
        # Set column widths for Servers sheet
        servers.column_dimensions['A'].width = 40  # Server Name
        servers.column_dimensions['B'].width = 25  # Serial Number
        
        for col, header in enumerate(servers_headers, 1):
            cell = servers.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            
        # Add sample server data
        for idx, server_info in enumerate(server_options, 2):
            parts = server_info.split(" | ")
            if len(parts) == 2:
                serial, name = parts
                servers.cell(row=idx, column=1, value=name)
                servers.cell(row=idx, column=2, value=serial)
        
        # Apply consistent styling to all worksheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Apply text wrapping and alignment to all cells
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
            
            # Freeze the header row in each sheet
            sheet.freeze_panes = 'A2'
        
        # Save the workbook
        wb.save(excel_file)
        print(f"Created standard Excel template: {excel_file}")
        print("\nImportant Note:")
        print("This template uses standard Excel dropdowns without dynamic filtering.")
        print("All servers will appear in every dropdown regardless of resource group.")
        print("To update dropdowns with real Intersight data, run:")
        print(f"  python3 update_intersight_data.py {excel_file}")
        
        return True
        
    except Exception as e:
        print(f"Error creating Excel template: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    excel_file = "output/Create_Intersight_Template.xlsx"
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    
    # Check if file exists and confirm replacement
    if os.path.exists(excel_file):
        response = input(f"A file named '{excel_file}' already exists. Would you like to keep the existing file or replace it? (k = keep, r = replace): ")
        if response.lower() == 'k':
            print(f"Keeping the existing file: {excel_file}.")
            sys.exit(0)
        elif response.lower() == 'r':
            print(f"Replacing the existing file: {excel_file}.")
        else:
            print("Invalid response. Exiting without changes.")
            sys.exit(1)
            
    create_standard_excel(excel_file)
