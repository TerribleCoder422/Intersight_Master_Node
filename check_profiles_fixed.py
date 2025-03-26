#!/usr/bin/env python3
from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('/Users/isaiahlaughlin/Lumos Dropbox/isaiah laughlin/Mac/Desktop/Python journey/Automate-Intersight/output/Intersight_Foundation.xlsx')

# Check Profiles sheet
profiles = wb['Profiles']
print(f'Number of rows in Profiles sheet: {profiles.max_row}')
print('Profile names and details:')
print('Row | Profile Name | Organization | Template | Server')
print('-' * 70)
for row in range(2, profiles.max_row+1):
    profile_name = profiles.cell(row=row, column=1).value
    org = profiles.cell(row=row, column=3).value  # Organization should be in column C
    template = profiles.cell(row=row, column=4).value  # Template should be in column D
    server = profiles.cell(row=row, column=5).value  # Server should be in column E
    print(f'{row} | {profile_name} | {org} | {template} | {server}')

# Check if data validations exist
print('\nData Validations:')
for sheet_name in ['Profiles', 'Policies', 'Template']:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        has_validations = hasattr(sheet, 'data_validations') and sheet.data_validations
        print(f'{sheet_name} sheet has data validations: {has_validations}')

# Check data validation formula content (lists of values)
print('\nDropdown Contents (if available):')
for sheet_name in ['Profiles', 'Policies', 'Template']:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if hasattr(sheet, 'data_validations') and sheet.data_validations:
            for dv in sheet.data_validations.dataValidation:
                if hasattr(dv, 'formula1') and dv.formula1:
                    range_str = str(dv.sqref) if hasattr(dv, 'sqref') else "Unknown"
                    print(f'{sheet_name}: {range_str} - Values: {dv.formula1}')
