#!/usr/bin/env python3
"""
Push Intersight Template Configuration to Cisco Intersight.
This script reads Excel template data and deploys the configuration to Intersight,
creating pools, policies, templates and profiles as specified.
"""

import pandas as pd
import os
import json
import traceback
import intersight
import requests
import time
import base64
import math
import hashlib
import hmac
import urllib.parse
import logging
import argparse
import sys
from tqdm import tqdm
from colorama import Fore, Style, init
from intersight.api_client import ApiClient
from intersight.configuration import Configuration
from intersight.rest import RESTResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import uuid
from datetime import datetime
import functools
import concurrent.futures
from typing import Dict, List, Any, Tuple, Optional

# Initialize colorama for colored terminal output
init(autoreset=True)

# Set up logging
log_dir = "logs"
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f"intersight_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

# Maximum number of retries for API calls
MAX_RETRIES = 3
# Delay between retries in seconds
RETRY_DELAY = 2
from intersight.api import (
    bios_api,
    boot_api,
    compute_api,
    fabric_api,
    firmware_api,
    ippool_api,
    macpool_api,
    organization_api,
    resource_api,
    server_api,
    storage_api,
    uuidpool_api,
    vnic_api
)

from dotenv import load_dotenv
import os

# Load environment variables from .env file
load_dotenv()

# Cache for API results
API_CACHE = {}

# Current version of the template
TEMPLATE_VERSION = "1.0.0"

# Global dictionary to store template name mappings
template_mappings = {}

def add_template_mapping(original_name, unique_name):
    """
    Add a mapping between original template name and unique template name
    """
    template_mappings[original_name] = unique_name
    print(f"Added template mapping: {original_name} -> {unique_name}")


def auto_adjust_column_width(worksheet, min_width=15, padding=2, custom_width_map=None, sheet_name=None):
    """
    Automatically adjust column widths based on content.
    
    Args:
        worksheet: The worksheet to adjust
        min_width: Minimum width for all columns
        padding: Extra characters to add for padding
        custom_width_map: Dictionary of column letters and their minimum widths
        sheet_name: Optional name of the sheet for logging purposes
    """
    if custom_width_map is None:
        custom_width_map = {}
    
    # Get actual sheet name if not provided
    if not sheet_name and hasattr(worksheet, 'title'):
        sheet_name = worksheet.title
    
    adjusted_columns = []
    for column in worksheet.columns:
        col_letter = get_column_letter(column[0].column)
        # Calculate max length of content in the column
        max_length = 0
        for cell in column:
            if cell.value:
                # Handle different data types
                if isinstance(cell.value, (int, float)):
                    cell_len = len(str(cell.value))
                elif isinstance(cell.value, str):
                    cell_len = len(cell.value)
                elif isinstance(cell.value, (datetime.datetime, datetime.date)):
                    cell_len = len(cell.value.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    cell_len = len(str(cell.value))
                max_length = max(max_length, cell_len)
        
        # Get minimum width from custom map or use default
        col_min_width = custom_width_map.get(col_letter, min_width)
        
        # Set the column width (content length + padding, but at least min_width)
        adjusted_width = max(max_length + padding, col_min_width)
        worksheet.column_dimensions[col_letter].width = adjusted_width
        adjusted_columns.append(col_letter)
    
    # Output message about auto-formatted columns
    if adjusted_columns:
        if sheet_name:
            print(f"Auto-adjusted column widths in {sheet_name} sheet: {', '.join(adjusted_columns)}")
        else:
            print(f"Auto-adjusted column widths: {', '.join(adjusted_columns)}")
    
    return adjusted_columns

# Configuration presets for common deployments
CONFIGURATION_PRESETS = {
    "VMware ESXi": {
        "BIOS": "VMware-Recommended",
        "BOOT": "ESXi-Boot",
        "vNIC": "VMware-Dual-vNIC",
        "QoS": "VMware-QoS",
        "Storage": "ESXi-Storage"
    },
    "Windows Server": {
        "BIOS": "Windows-Performance",
        "BOOT": "Windows-Boot",
        "vNIC": "Windows-Dual-vNIC",
        "QoS": "Windows-QoS",
        "Storage": "Windows-Storage"
    },
    "RedHat Enterprise Linux": {
        "BIOS": "RHEL-Performance",
        "BOOT": "RHEL-Boot",
        "vNIC": "RHEL-Dual-vNIC",
        "QoS": "RHEL-QoS",
        "Storage": "RHEL-Storage"
    }
}

# Known policy dependencies
POLICY_DEPENDENCIES = {
    "vNIC": ["QoS"],
    "LAN Connectivity": ["vNIC", "QoS"],
    "BOOT": ["Storage", "LAN Connectivity"],
    "Server Profile Template": ["BIOS", "BOOT", "LAN Connectivity", "Storage"]
}

def cached_api_call(timeout_minutes=5):
    """Decorator to cache API results with timeout to reduce calls to Intersight API."""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            # Create a cache key based on function name and arguments
            key = f"{func.__name__}:{str(args)}:{str(kwargs)}"
            now = datetime.datetime.now()
            
            # Check if result is in cache and not expired
            if key in API_CACHE and (now - API_CACHE[key]['timestamp']).total_seconds() < timeout_minutes * 60:
                print(f"Using cached result for {func.__name__}")
                return API_CACHE[key]['data']
            
            # Call the actual function
            try:
                result = func(*args, **kwargs)
            except Exception as e:
                print(f"Error updating Excel file with Intersight info: {str(e)}")
                import traceback
                traceback.print_exc()
                return False
            
            # Store result in cache
            API_CACHE[key] = {'data': result, 'timestamp': now}
            return result
        return wrapper
    return decorator
from intersight.model.vnic_lan_connectivity_policy import VnicLanConnectivityPolicy
from intersight.model.vnic_eth_if import VnicEthIf
from intersight.model.vnic_eth_adapter_policy import VnicEthAdapterPolicy
from intersight.model.vnic_eth_qos_policy import VnicEthQosPolicy
from intersight.model.fabric_eth_network_group_policy import FabricEthNetworkGroupPolicy
from intersight.model.server_profile_template import ServerProfileTemplate
import time
import argparse
import sys
import re
from copy import copy

def get_api_client():
    """
    Create an Intersight API client using the API key file
    """
    try:
        # Get API key details from environment variables
        api_key_id = os.getenv('INTERSIGHT_API_KEY_ID')
        api_key_file = os.getenv('INTERSIGHT_PRIVATE_KEY_FILE', './SecretKey.txt')
        
        if not api_key_id or not os.path.exists(api_key_file):
            print("Error: API key configuration not found")
            return None
            
        # Create configuration
        config = Configuration(
            host = os.getenv('INTERSIGHT_BASE_URL', 'https://intersight.com'),
            signing_info = intersight.signing.HttpSigningConfiguration(
                key_id = api_key_id,
                private_key_path = api_key_file,
                signing_scheme = intersight.signing.SCHEME_HS2019,
                signing_algorithm = intersight.signing.ALGORITHM_ECDSA_MODE_FIPS_186_3,
                hash_algorithm = intersight.signing.HASH_SHA256,
                signed_headers = [
                    intersight.signing.HEADER_REQUEST_TARGET,
                    intersight.signing.HEADER_HOST,
                    intersight.signing.HEADER_DATE,
                    intersight.signing.HEADER_DIGEST,
                ]
            )
        )
        
        # Create API client
        api_client = ApiClient(configuration=config)
        return api_client
        
    except Exception as e:
        print(f"Error creating API client: {str(e)}")
        return None

@cached_api_call(timeout_minutes=10)
def get_organizations(api_client):
    """
    Get list of organizations from Intersight
    """
    try:
        # Create API instance for organizations
        api_instance = organization_api.OrganizationApi(api_client)
        
        # Get list of organizations
        orgs = api_instance.get_organization_organization_list()
        
        # Extract organization names
        org_names = [org['Name'] for org in orgs.results]
        return org_names
    except Exception as e:
        print(f"Error getting organizations: {str(e)}")
        return ['default']  # Return default as fallback

@cached_api_call(timeout_minutes=10)
def get_organizations(api_client):
    """
    Get list of organizations from Intersight
    """
    if not api_client:
        print("Debug: No API client available, defaulting to 'default' organization")
        return ["default"]
        
    try:
        # Import here to avoid circular imports
        from intersight.api import organization_api, compute_api, server_api, vnic_api, bios_api, boot_api, storage_api, resource_api as resource_api_module_client
        print("Debug: Successfully created organization API client")
        
        orgs = org_api.get_organization_organization_list()
        print(f"Debug: Found organizations: {[org.name for org in orgs.results]}")
        
        return [org.name for org in orgs.results] if orgs.results else ["default"]
    except Exception as e:
        print(f"Debug: Error fetching organizations: {str(e)}")
        return ["default"]

def create_mac_pool(api_client, pool_data):
    """
    Create a MAC Pool in Intersight
    """
    from intersight.api import macpool_api
    from intersight.model.macpool_pool import MacpoolPool
    from intersight.model.macpool_block import MacpoolBlock
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, "Gruve")
        if not org_moid:
            print("Error: Gruve organization not found")
            return False

        # Use organization API to fetch organizations
        org_api = organization_api.OrganizationApi(api_client)
        orgs = org_api.get_organization_organization_list()
        org_names = [org.name for org in orgs.results]
        
        # Use compute API to fetch servers
        compute_api_instance = compute_api.ComputeApi(api_client)
        servers_list = compute_api_instance.get_compute_rack_unit_list()
        servers = servers_list.results
        
        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )

        # Create MAC pool block
        block = MacpoolBlock(
            class_id="macpool.Block",
            object_type="macpool.Block",
            _from=pool_data['_start_address'],  # Use our normalized address field
            size=int(pool_data['Size'])
        )
        
        # Create MAC pool
        pool = MacpoolPool(
            class_id="macpool.Pool",
            object_type="macpool.Pool",
            name=pool_data['Pool Name'],
            description=pool_data['Description'] if pd.notna(pool_data['Description']) else "",
            organization=org_ref,
            assignment_order="sequential",
            MacBlocks=[block]
        )
        
        # Create API instance
        api_instance = macpool_api.MacpoolApi(api_client)
        result = api_instance.create_macpool_pool(macpool_pool=pool)
        print(f"Successfully created MAC Pool: {result.name}")
        return True
        
    except Exception as e:
        print(f"Error creating MAC Pool: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_uuid_pool(api_client, pool_data):
    """
    Create a UUID Pool in Intersight
    """
    from intersight.api import uuidpool_api
    from intersight.model.uuidpool_pool import UuidpoolPool
    from intersight.model.uuidpool_uuid_block import UuidpoolUuidBlock
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, "Gruve")
        if not org_moid:
            print("Error: Gruve organization not found")
            return False

        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )

        # Create UUID pool block
        block = UuidpoolUuidBlock(
            class_id="uuidpool.UuidBlock",
            object_type="uuidpool.UuidBlock",
            _from=pool_data['_start_address'],  # Use our normalized address field
            size=int(pool_data['Size'])
        )
        
        # Create UUID pool
        pool = UuidpoolPool(
            class_id="uuidpool.Pool",
            object_type="uuidpool.Pool",
            name=pool_data['Pool Name'],
            description=pool_data['Description'] if pd.notna(pool_data['Description']) else "",
            organization=org_ref,
            assignment_order="sequential",
            prefix="000025B5-0000-0000",  # Standard prefix for UUIDs
            UuidSuffixBlocks=[block]
        )
        
        # Create API instance
        api_instance = uuidpool_api.UuidpoolApi(api_client)
        result = api_instance.create_uuidpool_pool(uuidpool_pool=pool)
        print(f"Successfully created UUID Pool: {result.name}")
        return True
        
    except Exception as e:
        print(f"Error creating UUID Pool: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def format_uuid_suffix(uuid_str):
    """Format a UUID suffix to match Intersight's expected pattern: XXXX-XXXXXXXXXXXX"""
    # Remove any non-hex characters and pad to 16 characters
    clean_uuid = ''.join(c for c in uuid_str if c.isalnum()).zfill(16)
    return f"{clean_uuid[:4]}-{clean_uuid[4:]}"

def pool_exists(api_client, pool_type, pool_name):
    """
    Check if a pool already exists in Intersight
    """
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, "Gruve")
        if not org_moid:
            return False

        # Create API instance based on pool type
        if pool_type == 'MAC Pool':
            from intersight.api import macpool_api
            api_instance = macpool_api.MacpoolApi(api_client)
            filter_str = f"Name eq '{pool_name}' and Organization.Moid eq '{org_moid}'"
            api_response = api_instance.get_macpool_pool_list(filter=filter_str)
        elif pool_type == 'UUID Pool':
            from intersight.api import uuidpool_api
            api_instance = uuidpool_api.UuidpoolApi(api_client)
            filter_str = f"Name eq '{pool_name}' and Organization.Moid eq '{org_moid}'"
            api_response = api_instance.get_uuidpool_pool_list(filter=filter_str)
        else:
            print(f"Unsupported pool type: {pool_type}")
            return False

        # Check if any pools were found
        return len(api_response.results) > 0

    except Exception as e:
        print(f"Error checking if pool exists: {str(e)}")
        return False

def create_pool(api_client, pool_data):
    """
    Create a pool in Intersight based on pool type
    """
    try:
        pool_type = pool_data['Pool Type']
        pool_name = pool_data['Pool Name']
        
        # Check if pool already exists
        if pool_exists(api_client, pool_type, pool_name):
            print(f"\n✅ DUPLICATE AVOIDED: Pool {pool_name} already exists in Intersight")
            print(f"Skipping creation to prevent duplicates")
            return True
            
        # Handle both 'Start Address' and 'First Address' column names
        start_address = None
        if 'Start Address' in pool_data and not pd.isna(pool_data['Start Address']):
            start_address = pool_data['Start Address']
        elif 'First Address' in pool_data and not pd.isna(pool_data['First Address']):
            start_address = pool_data['First Address']
        
        # Store the address for use by other functions
        pool_data['_start_address'] = start_address
            
        print(f"\nCreating {pool_type}: {pool_name}")
        print(f"Description: {pool_data['Description'] if pd.notna(pool_data['Description']) else 'None'}")
        print(f"Start/First Address: {start_address}")
        print(f"Size: {pool_data['Size']}")
        
        if pool_type == 'MAC Pool':
            return create_mac_pool(api_client, pool_data)
        elif pool_type == 'UUID Pool':
            return create_uuid_pool(api_client, pool_data)
        else:
            print(f"Unsupported pool type: {pool_type}")
            return False
            
    except Exception as e:
        print(f"Error creating pool: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def get_mac_pool_moid(api_client, pool_name, org_moid):
    """
    Get the MOID of a MAC Pool by name and organization MOID
    """
    from intersight.api import macpool_api
    
    api_instance = macpool_api.MacpoolApi(api_client)
    pools = api_instance.get_macpool_pool_list()
    for pool in pools.results:
        if pool.name == pool_name and pool.organization.moid == org_moid:
            return pool.moid
    return None

def get_pool_moid(api_client, pool_name):
    """
    Get the MOID of a pool by name
    """
    from intersight.api import macpool_api
    
    api_instance = macpool_api.MacpoolApi(api_client)
    pools = api_instance.get_macpool_pool_list(filter=f"Name eq '{pool_name}'").results
    
    if pools:
        return pools[0].moid
    else:
        raise Exception(f"Pool '{pool_name}' not found")

def get_policy_moid(api_client, policy_type, policy_name):
    """Get the MOID of a policy by name"""
    try:
        if policy_type == "bios.Policy":
            api_instance = bios_api.BiosApi(api_client)
            policies = api_instance.get_bios_policy_list()
        elif policy_type == "vnic.LanConnectivityPolicy":
            api_instance = vnic_api.VnicApi(api_client)
            policies = api_instance.get_vnic_lan_connectivity_policy_list()
        elif policy_type == "vnic.EthQosPolicy":
            api_instance = vnic_api.VnicApi(api_client)
            policies = api_instance.get_vnic_eth_qos_policy_list()
        elif policy_type == "storage.StoragePolicy":
            api_instance = storage_api.StorageApi(api_client)
            policies = api_instance.get_storage_storage_policy_list()
        elif policy_type == "macpool.Pool":
            api_instance = macpool_api.MacpoolApi(api_client)
            policies = api_instance.get_macpool_pool_list()
        elif policy_type == "boot.PrecisionPolicy":
            api_instance = boot_api.BootApi(api_client)
            policies = api_instance.get_boot_precision_policy_list()
        elif policy_type == "storage.StoragePolicies":
            api_instance = storage_api.StorageApi(api_client)
            policies = api_instance.get_storage_storage_policy_list()
        else:
            raise Exception(f"Unsupported policy type: {policy_type}")
        
        # Find the policy by name
        for policy in policies.results:
            if policy.name == policy_name:
                return policy.moid
                
        print(f"Policy {policy_name} not found")
        return None
        
    except Exception as e:
        print(f"Error getting policy MOID: {str(e)}")
        return None

def policy_exists(api_client, policy_type, policy_name):
    """
    Check if a policy already exists in Intersight
    """
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client)
        
        # Create API instance based on policy type
        if policy_type == "bios.Policy":
            api_instance = bios_api.BiosApi(api_client)
            response = api_instance.get_bios_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "vnic.EthQosPolicy":
            api_instance = vnic_api.VnicApi(api_client)
            response = api_instance.get_vnic_eth_qos_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "vnic.EthAdapterPolicy":
            api_instance = vnic_api.VnicApi(api_client)
            response = api_instance.get_vnic_eth_adapter_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "fabric.EthNetworkGroupPolicy":
            api_instance = fabric_api.FabricApi(api_client)
            response = api_instance.get_fabric_eth_network_group_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "vnic.LanConnectivityPolicy":
            api_instance = vnic_api.VnicApi(api_client)
            response = api_instance.get_vnic_lan_connectivity_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "boot.PrecisionPolicy":
            api_instance = boot_api.BootApi(api_client)
            response = api_instance.get_boot_precision_policy_list(filter=f"Name eq '{policy_name}'")
        elif policy_type == "storage.StoragePolicy":
            api_instance = storage_api.StorageApi(api_client)
            response = api_instance.get_storage_storage_policy_list(filter=f"Name eq '{policy_name}'")
        else:
            return False

        return len(response.results) > 0

    except Exception as e:
        print(f"Error checking if policy exists: {str(e)}")
        return False

def check_vnic_exists(api_client, vnic_name, lan_connectivity_moid):
    """
    Check if a vNIC already exists in the LAN Connectivity Policy
    """
    try:
        vnic_instance = vnic_api.VnicApi(api_client)
        vnic_list = vnic_instance.get_vnic_eth_if_list(filter=f"Name eq '{vnic_name}'")
        for vnic in vnic_list.results:
            if hasattr(vnic, 'lan_connectivity_policy') and vnic.lan_connectivity_policy.moid == lan_connectivity_moid:
                return True
        return False
    except Exception as e:
        print(f"Error checking vNIC existence: {str(e)}")
        return False

def move_sheet_after(workbook, sheet_to_move, target_sheet):
    """Move a worksheet to be right after another worksheet"""
    if target_sheet not in workbook.sheetnames or sheet_to_move not in workbook.sheetnames:
        return False
        
    # Get the indices
    target_index = workbook.sheetnames.index(target_sheet)
    current_index = workbook.sheetnames.index(sheet_to_move)
    
    # If sheet is already after target, do nothing
    if current_index == target_index + 1:
        return True
        
    # Remove and insert at new position
    sheet = workbook[sheet_to_move]
    workbook.remove(sheet)
    workbook.create_sheet(sheet_to_move, target_index + 1)
    
    # Copy the removed sheet to the new position
    new_sheet = workbook[sheet_to_move]
    for row in sheet.iter_rows():
        for cell in row:
            new_sheet[cell.coordinate].value = cell.value
            if cell.has_style:
                new_sheet[cell.coordinate].font = copy(cell.font)
                new_sheet[cell.coordinate].border = copy(cell.border)
                new_sheet[cell.coordinate].fill = copy(cell.fill)
                new_sheet[cell.coordinate].number_format = copy(cell.number_format)
                new_sheet[cell.coordinate].protection = copy(cell.protection)
                new_sheet[cell.coordinate].alignment = copy(cell.alignment)
    
    # Copy sheet properties
    new_sheet.sheet_format = copy(sheet.sheet_format)
    new_sheet.sheet_properties = copy(sheet.sheet_properties)
    new_sheet.merged_cells = copy(sheet.merged_cells)
    new_sheet.page_margins = copy(sheet.page_margins)
    new_sheet.page_setup = copy(sheet.page_setup)
    
    # Copy column dimensions
    for key, value in sheet.column_dimensions.items():
        new_sheet.column_dimensions[key] = copy(value)
    
    # Copy row dimensions
    for key, value in sheet.row_dimensions.items():
        new_sheet.row_dimensions[key] = copy(value)

def create_default_bios_policy(api_client, policy_name, org_name):
    """
    Create a default BIOS policy with standard settings
    """
    from intersight.api import bios_api
    from intersight.model.bios_policy import BiosPolicy
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False

        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Create BIOS policy with minimal defaults for AI workloads
        # Using only the most essential settings to avoid validation errors
        bios_policy = BiosPolicy(
            name=policy_name,
            description="Default BIOS Policy for AI workloads",
            organization=org_ref
            # Simplified settings to avoid validation errors
            # Additional settings can be added after creation through the UI
        )
        
        # Create API instance
        api_instance = bios_api.BiosApi(api_client)
        result = api_instance.create_bios_policy(bios_policy)
        print(f"Successfully created default BIOS Policy: {result.name}")
        return True
        
    except Exception as e:
        print(f"Error creating default BIOS policy: {str(e)}")
        return False

def create_default_boot_policy(api_client, policy_name, org_name):
    """
    Create a default Boot policy with standard settings
    """
    from intersight.api import boot_api
    from intersight.model.boot_precision_policy import BootPrecisionPolicy
    from intersight.model.boot_device_base import BootDeviceBase
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False

        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Create boot devices
        boot_local_disk = BootDeviceBase(
            class_id="boot.LocalDisk",
            object_type="boot.LocalDisk",
            name="LocalDisk",
            enabled=True,
            interface_type="sata",
            order=1
        )
        
        boot_pxe = BootDeviceBase(
            class_id="boot.Pxe",
            object_type="boot.Pxe",
            name="PXE",
            enabled=True,
            interface_source="name",
            ip_type="IPv4",
            interface_name="eth0",
            slot="MLOM",
            port=1,
            order=2
        )
        
        # Create boot policy with the boot devices
        boot_policy = BootPrecisionPolicy(
            name=policy_name,
            description=f"Default Boot Policy for AI workloads",
            organization=org_ref,
            configured_boot_mode="Uefi",
            boot_devices=[boot_local_disk, boot_pxe]
        )
        
        # Create API instance
        api_instance = boot_api.BootApi(api_client)
        result = api_instance.create_boot_precision_policy(boot_precision_policy=boot_policy)
        print(f"Successfully created default Boot Policy: {result.name}")
        return True
        
    except Exception as e:
        print(f"Error creating default Boot policy: {str(e)}")
        return False

def create_default_lan_policy(api_client, policy_name, org_name):
    """
    Create a default LAN Connectivity policy with standard settings
    """
    from intersight.api import vnic_api, macpool_api, fabric_api
    from intersight.model.vnic_lan_connectivity_policy import VnicLanConnectivityPolicy
    from intersight.model.vnic_eth_if import VnicEthIf
    from intersight.model.vnic_eth_adapter_policy import VnicEthAdapterPolicy
    from intersight.model.vnic_eth_qos_policy import VnicEthQosPolicy
    from intersight.model.fabric_eth_network_group_policy import FabricEthNetworkGroupPolicy  # Correct Network Group Policy
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False

        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Find MAC pools for both Fabric A and B
        mac_pool_instance = macpool_api.MacpoolApi(api_client)
        mac_pool_a_name = "AI_POD-MAC-A"
        mac_pool_b_name = "AI_POD-MAC-B"
        
        # Try to find MAC pool A by name
        mac_pools_a = mac_pool_instance.get_macpool_pool_list(filter=f"Name eq '{mac_pool_a_name}'")
        mac_pool_a_moid = None
        
        if mac_pools_a.results and len(mac_pools_a.results) > 0:
            mac_pool_a_moid = mac_pools_a.results[0].moid
            print(f"Using existing MAC pool A: {mac_pool_a_name} (MOID: {mac_pool_a_moid})")
        else:
            # If pool not found, try to find any available MAC pool
            all_mac_pools = mac_pool_instance.get_macpool_pool_list()
            if all_mac_pools.results and len(all_mac_pools.results) > 0:
                mac_pool_a_moid = all_mac_pools.results[0].moid
                mac_pool_a_name = all_mac_pools.results[0].name
                print(f"Using available MAC pool for Fabric A: {mac_pool_a_name} (MOID: {mac_pool_a_moid})")
            else:
                print("No MAC pools found. Cannot create LAN policy without MAC pool.")
                return False
        
        # Try to find MAC pool B by name
        mac_pools_b = mac_pool_instance.get_macpool_pool_list(filter=f"Name eq '{mac_pool_b_name}'")
        mac_pool_b_moid = None
        
        if mac_pools_b.results and len(mac_pools_b.results) > 0:
            mac_pool_b_moid = mac_pools_b.results[0].moid
            print(f"Using existing MAC pool B: {mac_pool_b_name} (MOID: {mac_pool_b_moid})")
        else:
            # If Fabric B pool not found, use the same as Fabric A
            mac_pool_b_moid = mac_pool_a_moid
            mac_pool_b_name = mac_pool_a_name
            print(f"Using MAC pool A for Fabric B: {mac_pool_b_name} (MOID: {mac_pool_b_moid})")
        
        # Create MAC pool references
        mac_pool_a_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="macpool.Pool",
            moid=mac_pool_a_moid
        )
        
        mac_pool_b_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="macpool.Pool",
            moid=mac_pool_b_moid
        )
        
        # Create LAN Connectivity policy first
        lan_policy = VnicLanConnectivityPolicy(
            name=policy_name,
            description="Default LAN Connectivity Policy for AI workloads",
            organization=org_ref,
            placement_mode="auto",
            target_platform="FIAttached"
        )
        
        # Create API instance
        api_instance = vnic_api.VnicApi(api_client)
        result = api_instance.create_vnic_lan_connectivity_policy(vnic_lan_connectivity_policy=lan_policy)
        print(f"Successfully created default LAN Connectivity Policy: {result.name}")
        
        # First create an ethernet adapter policy
        adapter_policy_name = f"{policy_name}_EthAdapter"
        adapter_policy = VnicEthAdapterPolicy(
            name=adapter_policy_name,
            description="Default Ethernet Adapter Policy",
            organization=org_ref,
            # Use simpler settings to avoid API compatibility issues
            rss_settings=True
            # The full interrupt settings weren't working properly
            # Removed both interrupt_settings and tcp_offload_settings
        )
        
        # Create the adapter policy
        adapter_result = api_instance.create_vnic_eth_adapter_policy(vnic_eth_adapter_policy=adapter_policy)
        print(f"Successfully created Ethernet Adapter Policy: {adapter_result.name}")
        
        # Create adapter policy reference
        adapter_policy_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="vnic.EthAdapterPolicy",
            moid=adapter_result.moid
        )
        
        # Create QoS policy (mandatory for vNICs)
        qos_policy_name = f"{policy_name}_QoS"
        qos_policy = VnicEthQosPolicy(
            name=qos_policy_name,
            description="Default Ethernet QoS Policy",
            organization=org_ref,
            mtu=1500,
            rate_limit=0,  # No rate limiting
            cos=0,  # Class of service
            trust_host_cos=False,
            burst=1024,  # Burst size in bytes
            priority="Best Effort"  # Valid options are Best Effort, FC, Platinum, Gold, Silver, Bronze
        )
        
        # Create the QoS policy
        qos_result = api_instance.create_vnic_eth_qos_policy(vnic_eth_qos_policy=qos_policy)
        print(f"Successfully created Ethernet QoS Policy: {qos_result.name}")
        
        # Create QoS policy reference
        qos_policy_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="vnic.EthQosPolicy",
            moid=qos_result.moid
        )
        
        # Create Network Group Policy (mandatory for vNICs) - this is different from Ethernet Network Policy
        network_group_policy_name = f"{policy_name}_NetworkGroup"
        network_group_policy = FabricEthNetworkGroupPolicy(
            name=network_group_policy_name,
            description="Default Ethernet Network Group Policy",
            organization=org_ref,
            vlan_settings={
                "native_vlan": 1,  # The native VLAN
                "allowed_vlans": "1"  # comma-separated list or hyphen-range of allowed VLANs
            }
        )
        
        # Create the Network Group policy using the appropriate API instance
        fabric_api_instance = fabric_api.FabricApi(api_client)
        network_group_result = fabric_api_instance.create_fabric_eth_network_group_policy(
            fabric_eth_network_group_policy=network_group_policy
        )
        print(f"Successfully created Ethernet Network Group Policy: {network_group_result.name}")
        
        # Create Network Group policy reference
        network_group_policy_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="fabric.EthNetworkGroupPolicy",
            moid=network_group_result.moid
        )
        
        # Create redundant vNIC interfaces for both Fabric A and B
        # vNIC for Fabric A
        vnic_eth_a = VnicEthIf(
            name="eth0",
            order=0,
            mac_pool=mac_pool_a_ref,  # Use our MAC pool A
            eth_adapter_policy=adapter_policy_ref,  # Associate with the Ethernet Adapter policy
            eth_qos_policy=qos_policy_ref,  # Associate with the QoS policy (mandatory)
            fabric_eth_network_group_policy=[network_group_policy_ref],  # Must be a list of MoMoRefs
            placement={
                "id": "MLOM",
                "pci_link": 0,
                "uplink": 0,
                "switch_id": "A"  # Connected to Fabric A
            },
            lan_connectivity_policy=MoMoRef(
                class_id="mo.MoRef",
                object_type="vnic.LanConnectivityPolicy",
                moid=result.moid
            )
        )
        
        # Create vNIC for Fabric B
        vnic_eth_b = VnicEthIf(
            name="eth1",
            order=1,
            mac_pool=mac_pool_b_ref,  # Use our MAC pool B
            eth_adapter_policy=adapter_policy_ref,  # Associate with the Ethernet Adapter policy
            eth_qos_policy=qos_policy_ref,  # Associate with the QoS policy (mandatory)
            fabric_eth_network_group_policy=[network_group_policy_ref],  # Must be a list of MoMoRefs
            placement={
                "id": "MLOM",
                "pci_link": 0,
                "uplink": 0,
                "switch_id": "B"  # Connected to Fabric B
            },
            lan_connectivity_policy=MoMoRef(
                class_id="mo.MoRef",
                object_type="vnic.LanConnectivityPolicy",
                moid=result.moid
            )
        )
        
        # Check if vNICs already exist before creating
        try:
            # Try to create vNIC for Fabric A
            vnic_a_result = api_instance.create_vnic_eth_if(vnic_eth_if=vnic_eth_a)
            print(f"Successfully created vNIC for Fabric A: {vnic_a_result.name}")
        except Exception as e:
            if "duplicate" in str(e).lower():
                print(f"vNIC 'eth0' already exists, using existing vNIC")
            else:
                print(f"Error creating vNIC for Fabric A: {str(e)}")
                # Continue with the process despite the error
        
        try:  
            # Try to create vNIC for Fabric B
            vnic_b_result = api_instance.create_vnic_eth_if(vnic_eth_if=vnic_eth_b)
            print(f"Successfully created vNIC for Fabric B: {vnic_b_result.name}")
        except Exception as e:
            if "duplicate" in str(e).lower():
                print(f"vNIC 'eth1' already exists, using existing vNIC")
            else:
                print(f"Error creating vNIC for Fabric B: {str(e)}")
                # Continue with the process despite the error
        
        return result.moid
        
    except Exception as e:
        print(f"Error creating default LAN Connectivity policy: {str(e)}")
        return False

def create_default_storage_policy(api_client, policy_name, org_name):
    """
    Create a default Storage policy with standard settings
    """
    from intersight.api import storage_api
    from intersight.model.storage_storage_policy import StorageStoragePolicy
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False

        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Create Storage policy with valid drive mode value
        # Cannot have both default_drive_mode="Jbod" and use_jbod_for_vd_creation=True
        storage_policy = StorageStoragePolicy(
            name=policy_name,
            description="Default Storage Policy for AI workloads",
            organization=org_ref,
            default_drive_mode="UnconfiguredGood",  # Must be one of ['UnconfiguredGood', 'Jbod', 'RAID0']
            use_jbod_for_vd_creation=False  # Cannot be True when in Jbod mode
        )
        
        # Create API instance
        api_instance = storage_api.StorageApi(api_client)
        result = api_instance.create_storage_storage_policy(storage_storage_policy=storage_policy)
        print(f"Successfully created default Storage Policy: {result.name}")
        return True
        
    except Exception as e:
        print(f"Error creating default Storage policy: {str(e)}")
        return False

def create_server_template(api_client, template_data):
    """
    Create a Server Profile Template in Intersight
    """
    from intersight.api import server_api
    from intersight.model.server_profile_template import ServerProfileTemplate
    from intersight.model.mo_mo_ref import MoMoRef
    import uuid
    
    try:
        # Handle different column formats - with or without asterisk
        if 'Template Name*' in template_data:
            template_name = template_data['Template Name*']
        else:
            template_name = template_data['Template Name']
            
        description = template_data.get('Description', '') if pd.notna(template_data.get('Description', '')) else ''
        
        if 'Organization*' in template_data:
            org_name = template_data['Organization*']
        else:
            org_name = template_data['Organization']
            
        if 'Target Platform*' in template_data:
            target_platform = template_data['Target Platform*']
        else:
            target_platform = template_data['Target Platform']
        
        # Get policy names from template data, marking them as required
        # For each policy type, check with/without asterisk (*) in column name
        
        # BIOS Policy (required)
        bios_policy = None
        if 'BIOS Policy*' in template_data and pd.notna(template_data['BIOS Policy*']):
            bios_policy = template_data['BIOS Policy*']
        elif 'BIOS Policy' in template_data and pd.notna(template_data['BIOS Policy']):
            bios_policy = template_data['BIOS Policy']
            
        # If no BIOS policy is specified, create a default one
        if not bios_policy:
            bios_policy = f"{template_name}_BIOS_Policy"
            print(f"⚠️ No BIOS policy specified. Creating default policy: {bios_policy}")
            create_default_bios_policy(api_client, bios_policy, org_name)
            
        # Boot Policy (required)
        boot_policy = None
        if 'Boot Policy*' in template_data and pd.notna(template_data['Boot Policy*']):
            boot_policy = template_data['Boot Policy*']
        elif 'Boot Policy' in template_data and pd.notna(template_data['Boot Policy']):
            boot_policy = template_data['Boot Policy']
            
        # If no Boot policy is specified, create a default one
        if not boot_policy:
            boot_policy = f"{template_name}_Boot_Policy"
            print(f"⚠️ No Boot policy specified. Creating default policy: {boot_policy}")
            create_default_boot_policy(api_client, boot_policy, org_name)
            
        # LAN Connectivity Policy (required)
        lan_policy = None
        if 'LAN Connectivity Policy*' in template_data and pd.notna(template_data['LAN Connectivity Policy*']):
            lan_policy = template_data['LAN Connectivity Policy*']
        elif 'LAN Connectivity Policy' in template_data and pd.notna(template_data['LAN Connectivity Policy']):
            lan_policy = template_data['LAN Connectivity Policy']
            
        # If no LAN policy is specified, create a default one
        if not lan_policy:
            lan_policy = f"{template_name}_LAN_Policy"
            print(f"⚠️ No LAN Connectivity policy specified. Creating default policy: {lan_policy}")
            create_default_lan_policy(api_client, lan_policy, org_name)
            
        # Storage Policy (required)
        storage_policy = None
        if 'Storage Policy*' in template_data and pd.notna(template_data['Storage Policy*']):
            storage_policy = template_data['Storage Policy*']
        elif 'Storage Policy' in template_data and pd.notna(template_data['Storage Policy']):
            storage_policy = template_data['Storage Policy']
            
        # If no Storage policy is specified, create a default one
        if not storage_policy:
            storage_policy = f"{template_name}_Storage_Policy"
            print(f"⚠️ No Storage policy specified. Creating default policy: {storage_policy}")
            create_default_storage_policy(api_client, storage_policy, org_name)
        
        print(f"Creating server template: {template_name}")
        print(f"Organization: {org_name}")
        print(f"Target Platform: {target_platform}")
        print(f"BIOS Policy: {bios_policy}")
        print(f"Boot Policy: {boot_policy}")
        print(f"LAN Policy: {lan_policy}")
        print(f"Storage Policy: {storage_policy}")
        
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False
        
        # Create a dictionary for the template
        template_dict = {}
        policy_bucket = []
        
        # Check if a template with this name already exists
        api_instance = server_api.ServerApi(api_client)
        # Try looking for exact template name first
        try:
            existing_templates = api_instance.get_server_profile_template_list(filter=f"Name eq '{template_name}'")
            if existing_templates.results and len(existing_templates.results) > 0:
                existing_template = existing_templates.results[0]
                print(f"Found existing template '{template_name}' with MOID {existing_template.moid}, reusing it")
                # Store the name mapping to make sure profiles can locate it
                add_template_mapping(template_name, existing_template.name)
                return True
        except Exception as e:
            print(f"Error checking for existing template: {str(e)}")
            # Continue with creating a new template
            
        # No existing template found, create a new one
        # Add the basic properties
        template_dict['name'] = template_name  # Use the exact name from Excel, no random suffix
        template_dict['description'] = description
        template_dict['target_platform'] = target_platform
        
        # Add organization reference
        template_dict['organization'] = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Add optional policies if specified
        # BIOS Policy (optional)
        if bios_policy:
            bios_policy_moid = get_policy_moid(api_client, "bios.Policy", bios_policy)
            if bios_policy_moid:
                policy_bucket.append(MoMoRef(
                    class_id="mo.MoRef",
                    object_type="bios.Policy",
                    moid=bios_policy_moid
                ))
            else:
                print(f"Warning: BIOS policy '{bios_policy}' not found, continuing without it")
        else:
            print("No BIOS policy specified, continuing without it")
        
        # LAN Connectivity Policy (optional)
        if lan_policy:
            lan_policy_moid = get_policy_moid(api_client, "vnic.LanConnectivityPolicy", lan_policy)
            if lan_policy_moid:
                policy_bucket.append(MoMoRef(
                    class_id="mo.MoRef",
                    object_type="vnic.LanConnectivityPolicy",
                    moid=lan_policy_moid
                ))
            else:
                print(f"Warning: LAN policy '{lan_policy}' not found, continuing without it")
        else:
            print("No LAN Connectivity policy specified, continuing without it")
        
        # Storage Policy (optional)
        if storage_policy:
            storage_policy_moid = get_policy_moid(api_client, "storage.StoragePolicy", storage_policy)
            if storage_policy_moid:
                policy_bucket.append(MoMoRef(
                    class_id="mo.MoRef",
                    object_type="storage.StoragePolicy",
                    moid=storage_policy_moid
                ))
            else:
                print(f"Warning: Storage policy '{storage_policy}' not found, continuing without it")
        else:
            print("No Storage policy specified, continuing without it")
        
        # Add Boot Policy only if specified and found
        if boot_policy:
            boot_policy_moid = get_policy_moid(api_client, "boot.PrecisionPolicy", boot_policy)
            if boot_policy_moid:
                policy_bucket.append(MoMoRef(
                    class_id="mo.MoRef",
                    object_type="boot.PrecisionPolicy",
                    moid=boot_policy_moid
                ))
            else:
                print(f"Warning: Boot policy {boot_policy} not found, continuing without boot policy")
        
        # Add policy bucket to template dictionary
        template_dict['policy_bucket'] = policy_bucket
        
        # Create the server profile template using the dictionary
        template = ServerProfileTemplate(**template_dict)
        
        # Create the template in Intersight
        api_instance = server_api.ServerApi(api_client)
        result = api_instance.create_server_profile_template(server_profile_template=template)
        
        print(f"Successfully created Server Template: {result.name}")
        # Store the original name for future reference
        add_template_mapping(template_name, result.name)
        return True
        
    except Exception as e:
        print(f"Error creating Server Template: {str(e)}")
        traceback.print_exc()
        return False

def get_server_moid(api_client, server_name):
    """
    Get the MOID of a server by name or serial number with flexible matching
    """
    from intersight.api import compute_api
    
    if not server_name or pd.isna(server_name):
        return None
        
    server_name = str(server_name).strip()
    
    # Check if we received a combined format "SERIAL | NAME"
    if " | " in server_name:
        serial, name = server_name.split(" | ", 1)
        serial = serial.strip()
        name = name.strip()
    else:
        # Try treating entire string as name or serial
        serial = server_name
        name = server_name
    
    try:
        print(f"Finding server with name: {name} or serial: {serial}")
        
        # Get all managed servers
        api_instance = compute_api.ComputeApi(api_client)
        response = api_instance.get_compute_physical_summary_list(
            filter="ManagementMode eq 'IntersightStandalone' or ManagementMode eq 'UCSM' or ManagementMode eq 'Intersight'"
        )
        
        # Print available servers for debugging
        print("Available servers:")
        for server in response.results:
            print(f"  - {server.name} (Serial: {server.serial}, MOID: {server.moid})")
        
        # Try finding server by serial first
        for server in response.results:
            if server.serial and server.serial.lower() == serial.lower():
                print(f"Found server by exact serial match: {server.name} (MOID: {server.moid})")
                return server.moid
        
        # Then try by name
        for server in response.results:
            if server.name and server.name.lower() == name.lower():
                print(f"Found server by exact name match: {server.name} (MOID: {server.moid})")
                return server.moid
                
        # Try partial match on name
        for server in response.results:
            if server.name and name.lower() in server.name.lower():
                print(f"Found server by partial name match: {server.name} (MOID: {server.moid})")
                return server.moid
                
        # Try partial match on serial
        for server in response.results:
            if server.serial and serial.lower() in server.serial.lower():
                print(f"Found server by partial serial match: {server.serial} (MOID: {server.moid})")
                return server.moid
        
        print(f"❌ Server not found: {server_name}")
        return None
        
    except Exception as e:
        print(f"Error finding server: {str(e)}")
        return None

def create_server_profile(api_client, profile_data, template_name=None, server_name=None, deploy="No"):
    """
    Create a profile from template using the approach from Cisco sample code
    """
    from intersight.api import server_api
    from intersight.model.server_profile import ServerProfile
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Get profile name - use the name key if available, otherwise Profile Name
        # This handles both formats (from Excel or direct dictionary input)
        if 'name' in profile_data:
            profile_name = profile_data['name']
        elif 'Profile Name' in profile_data:
            profile_name = profile_data['Profile Name']
        else:
            profile_name = profile_data.get('Profile Name*', 'Unknown Profile')
        
        # Get description with proper NaN handling
        if 'description' in profile_data:
            description = profile_data['description']
        else:
            description = profile_data.get('Description', '') 
            if pd.isna(description):
                description = ''
        
        # Get organization name from various possible keys
        if 'org_name' in profile_data:
            org_name = profile_data['org_name']
        elif 'Organization' in profile_data:
            org_name = profile_data['Organization']
        else:
            org_name = profile_data.get('Organization*', 'default')
            if pd.isna(org_name):
                org_name = 'default'
        
        # Use the passed template_name parameter if available, otherwise try to get from profile_data
        if template_name is None or template_name == '':
            if 'Template Name' in profile_data:
                template_name = profile_data['Template Name']
            else:
                template_name = profile_data.get('Template Name*', None)
        
        # Use the passed server_name parameter if available, otherwise try to get from profile_data
        if server_name is None or server_name == '':
            # Properly handle NaN values for server info
            server_info = ''
            if 'Server' in profile_data and not pd.isna(profile_data['Server']):
                server_info = str(profile_data['Server']).strip()
            elif 'Server*' in profile_data and not pd.isna(profile_data['Server*']):
                server_info = str(profile_data['Server*']).strip()
        else:
            server_info = server_name
        
        # Convert deploy parameter to correct case-insensitive format
        deploy_flag = str(deploy).strip().lower() == "yes"
        
        # Print details with proper values
        print(f"Creating server profile: {profile_name}")
        print(f"Organization: {org_name}")
        print(f"Template: {template_name}")
        print(f"Server: {server_info}")
        print(f"Deploy: {deploy}")
        
        if deploy.lower() == "yes":
            print(f"Note: The profile will be created but must be deployed manually in the Intersight UI.")
        
        # Get organization MOID
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print(f"Error: Organization {org_name} not found")
            return False
        
        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Get template MOID - Check if it's in our mappings first
        if template_name in template_mappings:
            mapped_name = template_mappings[template_name]
            print(f"Found template mapping for {template_name} -> {mapped_name}")
            template_moid = get_template_moid(api_client, mapped_name)
        else:
            template_moid = get_template_moid(api_client, template_name)
        
        if not template_moid:
            print(f"Error: Template {template_name} not found")
            return False
            
        # Create template reference
        template_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="server.ProfileTemplate",
            moid=template_moid
        )
        
        # Parse server info (if provided)
        server_moid = None
        if server_info:
            server_name = server_info.split(' | ')[0] if ' | ' in server_info else server_info
            server_moid = get_server_moid(api_client, server_name)
            if not server_moid:
                print(f"Error: Server {server_name} not found")
                return False
        
        # Create server reference if server_moid exists
        server_ref = None
        if server_moid:
            server_ref = MoMoRef(
                class_id="mo.MoRef",
                object_type="compute.RackUnit",
                moid=server_moid
            )

        try:
            from intersight.model.server_profile import ServerProfile
            from intersight.model.compute_physical_relationship import ComputePhysicalRelationship
            from intersight.model.organization_organization_relationship import OrganizationOrganizationRelationship
            from intersight.api import server_api
            
            # Create API instance
            api_instance = server_api.ServerApi(api_client)
            
            # Step 1: Create the profile using the SDK model objects (following Cisco sample)
            try:
                # Create the server profile from scratch
                server_profile = ServerProfile()
                server_profile.name = str(profile_name)
                server_profile.type = "instance"
                
                # Add organization reference with proper model object
                server_profile.organization = OrganizationOrganizationRelationship(
                    class_id="mo.MoRef",
                    object_type="organization.Organization",
                    moid=str(org_moid)
                )
                
                # Add template reference
                server_profile.src_template = {
                    "class_id": "mo.MoRef",
                    "object_type": "server.ProfileTemplate",
                    "moid": str(template_moid)
                }
                
                # Add description if provided
                if description:
                    server_profile.description = description
                    
                print(f"Creating profile {profile_name} from template {template_name}...")
                
                # Create the profile
                try:
                    print(f"Creating profile {profile_name} from template {template_name}...")
                    created_profile = api_instance.create_server_profile(server_profile)
                    profile_moid = created_profile.moid
                    print(f"Successfully created Server Profile: {profile_name} (MOID: {profile_moid})")
                
                    # Verify if server was assigned by checking the created profile
                    if server_moid:
                        try:
                            # Wait briefly to make sure server assignment is reflected in API
                            time.sleep(2)  
                            created_profile = api_instance.get_server_profile_by_moid(profile_moid)
                            
                            if hasattr(created_profile, 'assigned_server') and created_profile.assigned_server:
                                print(f"✅ SERVER SUCCESSFULLY ASSIGNED: Server {server_name} assigned to profile {profile_name}")
                                print(f"Profile is ready for manual deployment in the Intersight UI")
                            else:
                                print(f"⚠️ SERVER NOT ASSIGNED: Server {server_name} could not be assigned to profile {profile_name}")
                                print(f"Please assign server {server_name} manually in the Intersight UI")
                        except Exception as verify_error:
                            print(f"Could not verify server assignment status: {str(verify_error)}")
                            print(f"Please check the Intersight UI to confirm if server was assigned")
                except Exception as e:
                    print(f"Error creating profile: {str(e)}")
                    if hasattr(e, 'body'):
                        print(f"API Error: {e.body}")
                    return False
                
                # If we have a server to assign, do it in a separate patch operation
                if server_moid and not pd.isna(server_moid):
                    try:
                        print(f"Attaching server {server_name} to profile {profile_name}...")
                        # Create the server assignment patch
                        server_attachment = ServerProfile()
                        server_attachment.assigned_server = ComputePhysicalRelationship(
                            class_id="mo.MoRef",
                            object_type="compute.RackUnit",
                            moid=str(server_moid)
                        )
                        
                        # Patch the profile to add the server assignment
                        api_instance.patch_server_profile(
                            moid=profile_moid,
                            server_profile=server_attachment
                        )
                        print(f"Successfully attached server to profile {profile_name}")
                        
                        # Verify if server was assigned by checking the updated profile
                        time.sleep(2)
                        updated_profile = api_instance.get_server_profile_by_moid(profile_moid)
                        if hasattr(updated_profile, 'assigned_server') and updated_profile.assigned_server:
                            print(f"✅ SERVER SUCCESSFULLY ASSIGNED: Server {server_name} assigned to profile {profile_name}")
                            print(f"Profile is ready for manual deployment in the Intersight UI")
                        else:
                            print(f"⚠️ SERVER NOT ASSIGNED: Server {server_name} could not be assigned to profile {profile_name}")
                            print(f"Please assign server {server_name} manually in the Intersight UI")
                    except Exception as server_error:
                        print(f"Error attaching server to profile: {str(server_error)}")
                        if hasattr(server_error, 'body'):
                            print(f"API Error: {server_error.body}")
                        print(f"Please assign server {server_name} manually in the Intersight UI")
                
                # Do not automatically deploy profiles
                if deploy_flag:
                    print(f"NOTE: Your requested deploy=Yes but we NEVER automatically deploy profiles")
                    print(f"Please deploy profile {profile_name} manually from the Intersight UI")
                
                # Successfully created the profile
                return True
                        
            except Exception as api_error:
                print(f"Error creating profile: {str(api_error)}")
                print("\n⚠️ Unable to create server profile from template due to Intersight API limitations.")
                print("This profile will need to be created manually in the Intersight UI.")
                
                # Store profile for manual creation report
                if 'profiles_for_manual_creation' not in globals():
                    global profiles_for_manual_creation
                    profiles_for_manual_creation = []
                
                # Store info needed for manual creation
                profile_info = {
                    'name': profile_name,
                    'template': template_name,
                    'organization': org_name,
                    'server': server_name,
                    'deploy': deploy
                }
                profiles_for_manual_creation.append(profile_info)
                
                return False
            
            return True
            
        except Exception as e:
            print(f"Error creating Server Profile: {str(e)}")
            traceback.print_exc()
            return False
    
    except Exception as e:
        print(f"Error creating Server Profile: {str(e)}")
        traceback.print_exc()
        return False

def get_template_moid(api_client, template_name):
    """
    Get the MOID of a server profile template by name with flexible matching
    """
    from intersight.api import server_api
    
    try:
        # Check if there's a mapping entry for this template name
        if template_name in template_mappings:
            mapped_name = template_mappings[template_name]
            print_info(f"Found template mapping for {template_name} -> {mapped_name}")
            template_name = mapped_name
        
        # Create API instance
        api_instance = server_api.ServerApi(api_client)
        
        # First try exact match
        filter_str = f"Name eq '{template_name}'"
        response = api_instance.get_server_profile_template_list(filter=filter_str)
        
        # Check if exact match template exists
        if response.results and len(response.results) > 0:
            print_success(f"Found exact match for template: {template_name}")
            return response.results[0].moid
            
        # If exact match not found, try case-insensitive search
        print_info(f"Exact match for '{template_name}' not found, trying flexible matching...")
        
        # Get all templates and search for a close match
        all_templates = api_instance.get_server_profile_template_list()
        
        # Flexible matching options (in order of preference):
        # 1. Exact match (already tried above)
        # 2. Case-insensitive exact match
        # 3. Template name starts with our search term
        # 4. Template name contains our search term
        template_matches = []
        
        if all_templates.results:
            template_name_lower = template_name.lower()
            
            for tmpl in all_templates.results:
                # Case-insensitive exact match
                if tmpl.name.lower() == template_name_lower:
                    print_success(f"Found case-insensitive match: {tmpl.name}")
                    return tmpl.moid
                
                # Template name starts with our search term
                if tmpl.name.lower().startswith(template_name_lower):
                    template_matches.append((1, tmpl))  # Priority 1
                    continue
                    
                # Template name contains our search term
                if template_name_lower in tmpl.name.lower():
                    template_matches.append((2, tmpl))  # Priority 2
            
            # Sort by priority (lower number is higher priority)
            template_matches.sort(key=lambda x: x[0])
            
            if template_matches:
                best_match = template_matches[0][1]
                print_success(f"Found best match for template '{template_name}': {best_match.name}")
                return best_match.moid
                
        print_error(f"No matching template found for '{template_name}'")
        return None
            
    except Exception as e:
        print(f"Error getting template MOID: {str(e)}")
        return None

def create_server_template_from_excel(api_client, excel_file):
    """
    Create Server Profile Templates from Excel
    """
    try:
        # Read Excel file
        print("\nCreating server templates from Excel...")
        df_sheets = pd.read_excel(excel_file, sheet_name=None)
        
        # Check if we have either Template or Templates sheet
        template_sheet = None
        if 'Template' in df_sheets:
            print("Found 'Template' sheet")
            template_sheet = 'Template'
        elif 'Templates' in df_sheets:
            print("Found 'Templates' sheet")
            template_sheet = 'Templates'
        
        if not template_sheet:
            print("Warning: Neither 'Template' nor 'Templates' sheet found in the Excel file")
            return False
        
        # Get the template dataframe
        df_templates = df_sheets[template_sheet]
        
        if df_templates.empty:
            print(f"No templates defined in {template_sheet} sheet.")
            return False
            
        # Print column information for debugging
        print(f"Template sheet columns: {df_templates.columns.tolist()}")
        
        # Process each row
        for index, row in df_templates.iterrows():
            # Skip rows with no template name
            if pd.isna(row['Template Name*']) if 'Template Name*' in df_templates.columns else pd.isna(row['Template Name']):
                continue
                
            # Extract template data using column names
            if 'Template Name*' in df_templates.columns:
                template_name = row['Template Name*']
            else:
                template_name = row['Template Name']
                
            if 'Organization*' in df_templates.columns:
                org_name = row['Organization*'] if not pd.isna(row['Organization*']) else "default"
            else:
                org_name = row['Organization'] if not pd.isna(row['Organization']) else "default"
                
            # Resource Group
            if 'Resource Group*' in df_templates.columns:
                resource_group = row['Resource Group*'] if not pd.isna(row['Resource Group*']) else None
            elif 'Resource Group' in df_templates.columns:
                resource_group = row['Resource Group'] if not pd.isna(row['Resource Group']) else None
            else:
                resource_group = None
                
            # Description
            if 'Description' in df_templates.columns:
                description = row['Description'] if not pd.isna(row['Description']) else ""
            else:
                description = ""
                
            # Target Platform
            if 'Target Platform*' in df_templates.columns:
                target_platform = row['Target Platform*'] if not pd.isna(row['Target Platform*']) else "FIAttached"
            elif 'Target Platform' in df_templates.columns:
                target_platform = row['Target Platform'] if not pd.isna(row['Target Platform']) else "FIAttached"
            else:
                target_platform = "FIAttached"
            
            print(f"  Creating template: {template_name} (Organization: {org_name}, Platform: {target_platform})")
            # Get organization MOID
            org_moid = get_org_moid(api_client, org_name)
            if not org_moid:
                print_warning(f"  - Could not find organization: {org_name}")
                continue
                
            # Create template data structure with the correct keys for create_server_template
            template_data = {}
            
            # Add Template Name with/without asterisk to support both functions
            template_data['Template Name'] = template_name
            template_data['Template Name*'] = template_name
            
            # Add Organization with/without asterisk
            template_data['Organization'] = org_name
            template_data['Organization*'] = org_name
            
            # Add Target Platform with/without asterisk
            template_data['Target Platform'] = target_platform
            template_data['Target Platform*'] = target_platform
            
            # Add Description
            template_data['Description'] = description
            
            # Add Resource Group
            if resource_group:
                template_data['Resource Group'] = resource_group
                template_data['Resource Group*'] = resource_group
            
            # Call existing template creation function
            create_success = create_server_template(api_client, template_data)
            
            if not create_success:
                print(f"Failed to create template: {template_name}")
            
        return True
    except Exception as e:
        print_error(f"Error creating server templates: {str(e)}")
        traceback.print_exc()
        return False

def create_server_profiles_from_excel(api_client, excel_file):
    """
    Read the Excel template and create server profiles in Intersight
    """
    print("\nCreating server profiles from Excel...")
    try:
        df_profiles = pd.read_excel(excel_file, sheet_name='Profiles')
        if df_profiles.empty:
            print_warning("No profiles defined in Profiles sheet.")
            return False
        
        # Print column headers for debugging
        print("Column headers in Profiles sheet:")
        print(df_profiles.columns.tolist())
        
        # Count total and deploy-marked profiles for reporting
        total_profiles = 0
        deploy_profiles = 0
            
        # Process all profiles in the sheet
        for index, row in df_profiles.iterrows():
            # Original code was skipping the first row incorrectly
            # Skip only if it's a header row with "Profile Name*" as cell content
            # or if the profile name cell is empty
            
            # Debug the row data
            print(f"Row {index}: {row.tolist()}")
            
            # If profile name is empty, skip
            profile_name_value = row['Profile Name*'] if 'Profile Name*' in row else row.iloc[0] if len(row) > 0 else None
            if pd.isna(profile_name_value):
                print(f"  Skipping row {index}: empty profile name")
                continue
                
            total_profiles += 1
            
            # Get values using column names instead of numeric indices
            profile_name = row['Profile Name*'] if 'Profile Name*' in row.index else row.iloc[0]
            description = row['Description'] if 'Description' in row.index and not pd.isna(row['Description']) else ""
            org_name = row['Organization*'] if 'Organization*' in row.index and not pd.isna(row['Organization*']) else "default"
            
            # Try to get resource group by name, fall back to position
            if 'Resource Group*' in row.index:
                resource_group = row['Resource Group*'] if not pd.isna(row['Resource Group*']) else None
            else:
                resource_group = row.iloc[3] if len(row) > 3 and not pd.isna(row.iloc[3]) else None
                
            # Get template name by column name or position
            if 'Template Name*' in row.index:
                template_name = row['Template Name*'] if not pd.isna(row['Template Name*']) else None
            else:
                template_name = row.iloc[4] if len(row) > 4 and not pd.isna(row.iloc[4]) else None
                
            # Get server name by column name or position
            if 'Server*' in row.index:
                server_name = row['Server*'] if not pd.isna(row['Server*']) else None
            else:
                server_name = row.iloc[5] if len(row) > 5 and not pd.isna(row.iloc[5]) else None
                
            # Get deploy flag by column name or position
            if 'Deploy*' in row.index:
                deploy = row['Deploy*'] if not pd.isna(row['Deploy*']) else "No"
            else:
                deploy = row.iloc[7] if len(row) > 7 and not pd.isna(row.iloc[7]) else "No"
            
            # Print debug info about each profile
            print(f"  Profile: {profile_name}, Deploy value: '{deploy}', Type: {type(deploy)}")
            
            # Only process profiles with Deploy set to Yes (case insensitive)
            deploy_str = str(deploy).strip().lower()
            if deploy_str != "yes":
                print(f"  Skipping profile: {profile_name} (Deploy value is '{deploy}', not 'Yes')")
                continue
                
            deploy_profiles += 1
            print(f"  Creating profile: {profile_name} (Template: {template_name}, Server: {server_name}, Organization: {org_name})")
            
            # Prepare profile data dictionary
            profile_data = {
                'Profile Name': profile_name,
                'Description': description,
                'Organization': org_name,
                'Resource Group': resource_group,
                'Template Name': template_name
            }
            
            # Convert deploy string to boolean-like string for API call
            deploy_value = "Yes" if deploy_str == "yes" else "No"
            print(f"  Setting deploy value to: {deploy_value}")
            
            # Create the profile
            create_server_profile(api_client, profile_data, template_name, server_name, deploy_value)
            
        # Print summary
        print(f"\nProfile Creation Summary:")
        print(f"  Total profiles in Excel: {total_profiles}")
        print(f"  Profiles marked for deployment: {deploy_profiles}")
        print(f"  Profiles attempted: {deploy_profiles}")
        
        # Print a message if any profiles need to be created manually
        if 'profiles_for_manual_creation' in globals() and profiles_for_manual_creation:
            print("\n" + "=" * 80)
            print("\n⚠️  PROFILES REQUIRING MANUAL CREATION IN INTERSIGHT  ⚠️\n")
            print("The following profiles must be created manually in the Intersight UI")
            print("due to API limitations when creating profiles from templates:\n")
            print("-" * 50)
            print(f"{'Profile Name':<20} {'Template':<30} {'Server':<25} {'Deploy':<5}")
            print("-" * 80)
            for profile in profiles_for_manual_creation:
                print(f"{profile['name']:<20} {profile['template']:<30} {profile['server'] if profile['server'] else 'nan':<25} {profile['deploy']:<5}")
            print("\n" + "-" * 50)
            print("Instructions:")
            print("1. Log into Intersight UI")
            print("2. Navigate to Profiles > UCS Server Profiles")
            print("3. Click 'Create Server Profile'")
            print("4. Select 'From Template'")
            print("5. For each profile above:")
            print("   - Select the listed template")
            print("   - Enter the profile name")
            print("   - Assign the server (if listed)")
            print("   - Deploy if 'Deploy' is set to 'Yes'")
            print("\n" + "=" * 80)
        
        return True
    except Exception as e:
        print_error(f"Error creating server profiles: {str(e)}")
        traceback.print_exc()
        return False
def process_foundation_template(excel_file):
    """
    Read the Excel template and create pools and policies in Intersight
    """
    try:
        # Read Excel file
        print_info("Reading Excel file...")
        df = pd.read_excel(excel_file, sheet_name=None)
        
        # Get API client
        print_info("Connecting to Intersight API...")
        api_client = get_api_client()
        if not api_client:
            print_error("Failed to get API client")
            return False
            
        # Process Pools sheet first
        if 'Pools' in df:
            pools_df = df['Pools']
            # Rename columns to remove asterisks
            pools_df.columns = pools_df.columns.str.replace('*', '')
            
            # Validate pools data before processing
            print_info("Validating pools data...")
            invalid_pools = validate_pools_data(pools_df)
            if invalid_pools:
                print_error(f"Found {len(invalid_pools)} invalid pool configurations:")
                for pool_info in invalid_pools:
                    print_error(f"  - {pool_info}")
                print_error("Please correct these issues in the Excel file before proceeding.")
                return False
            
            # Track pool creation success
            successful_pools = []
            failed_pools = []
            
            # Create or verify each pool with progress bar
            print_info("\nProcessing pools...")
            for _, row in progress_bar(pools_df.iterrows(), desc="Creating Pools", total=len(pools_df)):
                pool_name = row['Pool Name']
                pool_type = row['Pool Type']
                
                # Update progress bar description
                sys.stdout.write(f"\rProcessing {pool_name}...")
                sys.stdout.flush()
                
                # Check if pool exists
                if pool_exists(api_client, pool_type, pool_name):
                    print_info(f"Pool {pool_name} already exists, skipping creation")
                    successful_pools.append(f"{pool_name} (already exists)")
                    continue
                    
                # Try to create the pool
                if create_pool(api_client, row):
                    successful_pools.append(pool_name)
                else:
                    failed_pools.append(pool_name)
            
            # Print summary of pool creation
            print_summary("Pool Creation", successful_pools, failed_pools)
            
            # If any pools failed to create, stop here
            if failed_pools:
                print_error("\nAborting further processing until pool creation issues are resolved.")
                return False
                
            print_success("\nAll pools created or verified successfully.")
                
        # Only proceed with policies if pools were successful
        if 'Policies' in df:
            policies_df = df['Policies']
            # Rename columns to remove asterisks
            policies_df.columns = policies_df.columns.str.replace('*', '')
            
            # Validate policies data before processing
            print_info("Validating policies data...")
            invalid_policies = validate_policies_data(policies_df)
            if invalid_policies:
                print_error(f"Found {len(invalid_policies)} invalid policy configurations:")
                for policy_info in invalid_policies:
                    print_error(f"  - {policy_info}")
                print_error("Please correct these issues in the Excel file before proceeding.")
                return False
            
            # Create policies in order: BIOS, QoS, vNIC, Boot, Storage
            print_info("\nProcessing policies...")
            policy_order = ['BIOS', 'QoS', 'vNIC', 'Boot', 'Storage']
            
            # Track policy creation success
            successful_policies = []
            failed_policies = []
            
            for policy_type in policy_order:
                policy_rows = policies_df[policies_df['Policy Type'] == policy_type]
                if len(policy_rows) == 0:
                    continue
                    
                print_info(f"\nCreating {policy_type} policies...")
                for _, row in progress_bar(policy_rows.iterrows(), desc=f"Creating {policy_type} Policies", total=len(policy_rows)):
                    policy_name = row['Policy Name']
                    
                    # Update progress bar description
                    sys.stdout.write(f"\rProcessing {policy_name}...")
                    sys.stdout.flush()
                    
                    # Check if policy exists
                    if policy_exists(api_client, get_policy_class_id(policy_type), policy_name):
                        print(f"✅ DUPLICATE AVOIDED: Policy {policy_name} already exists in Intersight")
                        print(f"Skipping creation to prevent duplicates")
                        successful_policies.append(f"{policy_name} (already exists)")
                        continue
                        
                    # Try to create the policy
                    if create_policy(api_client, row):
                        successful_policies.append(f"{policy_type}: {policy_name}")
                    else:
                        failed_policies.append(f"{policy_type}: {policy_name}")
                        print_error(f"Failed to create policy {policy_name}")
                        break  # Stop processing this policy type if one fails
                
                # If any policies failed in this batch, stop processing
                if failed_policies:
                    print_error(f"\nError: Failed to create the following policies:")
                    for policy in failed_policies:
                        print_error(f"  - {policy}")
                    print_error(f"\nAborting further processing until policy creation issues are resolved.")
                    return False
                    
                print_success(f"All {policy_type} policies created or verified successfully.")
                
                # Add a small delay between policy types
                if policy_type != policy_order[-1]:
                    print_info(f"Waiting for {policy_type} policies to be fully created...")
                    time.sleep(5)
            
            print("\nAll policies created or verified successfully.")

        # Process Templates sheet
        if 'Template' in df:
            templates_df = df['Template']
            # Rename columns to remove asterisks
            templates_df.columns = templates_df.columns.str.replace('*', '')
            
            print("\nProcessing server templates...")
            
            # Track template creation success
            templates_created = True
            failed_templates = []
            
            for _, row in templates_df.iterrows():
                template_name = row['Template Name']
                print(f"\nCreating server template: {template_name}")
                
                # Create the template
                if not create_server_template(api_client, row):
                    templates_created = False
                    failed_templates.append(template_name)
            
            # If any templates failed, notify but continue processing
            if not templates_created:
                print("\nWarning: Failed to create the following templates:")
                for template in failed_templates:
                    print(f"  - {template}")
            else:
                print("\nAll server templates created or verified successfully.")
            
            # Add a small delay after template creation
            print("Waiting for templates to be fully created...")
            time.sleep(5)
        
        # Process Profiles sheet
        if 'Profiles' in df:
            profiles_df = df['Profiles']
            # Rename columns to remove asterisks
            profiles_df.columns = profiles_df.columns.str.replace('*', '')
            
            print("\nProcessing server profiles...")
            
            # Track profile creation success
            profiles_created = True
            failed_profiles = []
            
            for _, row in profiles_df.iterrows():
                profile_name = row['Profile Name']
                print(f"\nCreating server profile: {profile_name}")
                
                # Check if profile should be deployed
                deploy = row.get('Deploy', 'No')
                if deploy.lower() == 'yes':
                    print(f"Profile {profile_name} will be deployed after creation")
                
                # Create the profile using the new approach that derives from template
                if not create_and_derive_profile(api_client, row):
                    profiles_created = False
                    failed_profiles.append(profile_name)
            
            # If any profiles failed, notify
            if not profiles_created:
                print("\nWarning: Failed to create the following profiles:")
                for profile in failed_profiles:
                    print(f"  - {profile}")
            else:
                print("\nAll server profiles created or verified successfully.")
            
        # Display a summary of profiles that need manual creation
        if 'profiles_for_manual_creation' in globals() and profiles_for_manual_creation:
            print("\n" + "="*80)
            print("\n⚠️  PROFILES REQUIRING MANUAL CREATION IN INTERSIGHT  ⚠️")
            print("\nThe following profiles must be created manually in the Intersight UI")
            print("due to API limitations when creating profiles from templates:")
            print("\n" + "-"*50)
            print(f"{'Profile Name':<20} {'Template':<30} {'Server':<25} {'Deploy':<5}")
            print("-"*80)
            
            for profile in profiles_for_manual_creation:
                name = profile.get('name', 'N/A')
                template = profile.get('template', 'N/A')
                server = profile.get('server', 'N/A')
                deploy = profile.get('deploy', 'No')
                print(f"{name:<20} {template:<30} {server:<25} {deploy:<5}")
                
            print("\n" + "-"*50)
            print("Instructions:")
            print("1. Log into Intersight UI")
            print("2. Navigate to Profiles > UCS Server Profiles")
            print("3. Click 'Create Server Profile'")
            print("4. Select 'From Template'")
            print("5. For each profile above:")
            print("   - Select the listed template")
            print("   - Enter the profile name")
            print("   - Assign the server (if listed)")
            print("   - Deploy if 'Deploy' is set to 'Yes'")
            print("\n" + "="*80)
        
        print("\nCompleted processing the Foundation template")
        return True
        
    except Exception as e:
        print(f"\nError processing Foundation template: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def create_and_push_configuration(api_client, excel_file):
    """
    Read the Excel template and create pools and policies in Intersight
    """
    try:
        # Read Excel file
        df = pd.read_excel(excel_file, sheet_name=None)
        
        # Process Pools sheet
        if 'Pools' in df:
            pools_df = df['Pools']
            for _, row in pools_df.iterrows():
                create_pool(api_client, row)
                
        # Process Policies sheet in specific order
        if 'Policies' in df:
            policies_df = df['Policies']
            
            # Create policies in order: BIOS, QoS, vNIC, Boot, Storage
            policy_order = ['BIOS', 'QoS', 'vNIC', 'Boot', 'Storage']
            
            for policy_type in policy_order:
                policy_rows = policies_df[policies_df['Policy Type'] == policy_type]
                for _, row in policy_rows.iterrows():
                    if policy_exists(api_client, get_policy_class_id(policy_type), row['Name']):
                        print(f"Policy {row['Name']} already exists, skipping creation")
                    else:
                        create_policy(api_client, row)
                    
        print("Completed processing the Foundation template")
        return True
        
    except Exception as e:
        print(f"Error processing Foundation template: {str(e)}")
        return False

def get_policy_class_id(policy_type):
    """Get the class ID for a policy type"""
    policy_map = {
        'BIOS': 'bios.Policy',
        'QoS': 'vnic.EthQosPolicy',
        'vNIC': 'vnic.LanConnectivityPolicy',
        'Storage': 'storage.StoragePolicy',
        'Boot': 'boot.PrecisionPolicy'
    }
    return policy_map.get(policy_type)

def add_template_sheet(excel_file, api_client):
    """Add or update the Template sheet with dropdowns"""
    try:
        # Load workbook
        workbook = load_workbook(excel_file)
        
        # Store existing values if sheet exists
        existing_values = []
        if 'Template' in workbook.sheetnames:
            template_sheet = workbook['Template']
            existing_values = list(template_sheet.iter_rows(min_row=2, values_only=True))
            workbook.remove(template_sheet)
        
        # Create new sheet
        template_sheet = workbook.create_sheet(title='Template')
        
        # Add headers
        headers = [
            "Template Name*", 
            "Organization*", 
            "Description",
            "Target Platform*",
            "BIOS Policy*",
            "Boot Policy*",
            "LAN Connectivity Policy*",
            "Storage Policy*"
        ]
        
        # Define styles
        header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
        return True
        
    except Exception as e:
        print(f"Error adding template sheet: {str(e)}")
        return False
        
def create_template_excel(excel_file):
    """Create a fresh template Excel file with the original structure"""
    # Check if file exists, prompt to keep or replace
    if os.path.exists(excel_file):
        response = input(f"A file named '{excel_file}' already exists. Would you like to keep the existing file or replace it? (k = keep, r = replace): ")
        if response.lower() == 'k':
            print(f"Keeping the existing file: {excel_file}.")
            return True
        elif response.lower() != 'r':
            print("Invalid response. Exiting.")
            return False
            
    workbook = Workbook()
    
    # Remove default sheet
    default = workbook.active
    workbook.remove(default)
    
    # Create sheets in the correct order
    sheets = [
        'Pools',
        'Policies',
        'Template',
        'Profiles',
        'Templates',  # Info sheet
        'Organizations',  # Info sheet
        'Servers'  # Info sheet
    ]
    
    # Create sample lists for dropdowns - these will be populated from Intersight when the automation runs
    org_list = ["default", "DevOps", "Production", "Test", "UAT"]
    server_list = ["Server-1 (FCH1234V5Z7)", "Server-2 (FCH5678A9BC)", "Server-3 (FCH9012D3EF)"]
    
    # Create all sheets first
    for sheet_name in sheets:
        workbook.create_sheet(sheet_name)
    
    # Define styles - using a lighter shade of green for a more subtle look
    header_fill = PatternFill(start_color='A0D7BE', end_color='A0D7BE', fill_type='solid')  # Light green
    header_font = Font(color='000000', bold=True)  # Black text for readability
    
    # Set up Pools sheet
    pools_sheet = workbook['Pools']
    headers = ["Pool Type*", "Pool Name*", "Description", "Start Address*", "Size*"]
    
    for col, header in enumerate(headers, 1):
        # Apply the header styling
        cell = pools_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        # All headers should be black, even those with asterisks
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample pool data
    sample_pools = [
        ("MAC Pool", "Ai_POD-MAC-A", "MAC Pool for AI POD Fabric A", "00:25:B5:A0:00:00", "256"),
        ("MAC Pool", "Ai_POD-MAC-B", "MAC Pool for AI POD Fabric B", "00:25:B5:B0:00:00", "256"),
        ("UUID Pool", "Ai_POD-UUID-Pool", "UUID Pool for AI POD Servers", "0000-000000000001", "100")
    ]
    
    # Add all sample pool data to the sheet
    for idx, example in enumerate(sample_pools, 2):
        for col, value in enumerate(example, 1):
            pools_sheet.cell(row=idx, column=col, value=value)
    
    # Set up Policies sheet
    policies_sheet = workbook['Policies']
    policies_headers = ["Policy Type*", "Policy Name*", "Description", "Organization*"]
    
    for col, header in enumerate(policies_headers, 1):
        # Apply the header styling
        cell = policies_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        # All headers should be black, even those with asterisks
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample policy data
    sample_policies = [
        ('vNIC', 'Ai_POD-vNIC-A', 'vNIC Policy for AI POD Fabric A', 'default'),
        ('vNIC', 'Ai_POD-vNIC-B', 'vNIC Policy for AI POD Fabric B', 'default'),
        ('BIOS', 'Ai_POD-BIOS', 'BIOS Policy for AI POD', 'default'),
        ('BOOT', 'Ai_POD-BOOT', 'Boot Policy for AI POD', 'default'),
        ('QoS', 'Ai_POD-QoS', 'QoS Policy for AI POD', 'default'),
        ('Storage', 'Ai_POD-Storage', 'Storage Policy for AI POD', 'default')
    ]
    for row_idx, row in enumerate(sample_policies, 2):
        for col_idx, value in enumerate(row, 1):
            policies_sheet.cell(row=row_idx, column=col_idx, value=value)
            
    # Add organization dropdown to column D in Policies sheet
    org_validation_policies = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
    policies_sheet.add_data_validation(org_validation_policies)
    org_validation_policies.add('D2:D1000')  # Column D
    
    # Set up Template sheet - use the existing sheet
    template_sheet = workbook['Template']
    template_headers = [
        "Template Name*", 
        "Organization*", 
        "Resource Group*",
        "Description",
        "Target Platform*",
        "BIOS Policy*",
        "Boot Policy*",
        "LAN Connectivity Policy*",
        "Storage Policy*"
    ]
    for col, header in enumerate(template_headers, 1):
        # Split header for asterisk
        if header.endswith('*'):
            base = header.rstrip('*')
            cell = template_sheet.cell(row=1, column=col, value=base + '*')
        else:
            cell = template_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        # All headers should be black, even those with asterisks
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample template data
    template_example = [
        "Ai_POD_Template",
        "default",
        "AI POD Servers",
        "Server template for AI POD workloads",
        "FIAttached",
        "Ai_POD-BIOS",
        "Ai_POD-BOOT",
        "Ai_POD-vNIC-A",
        "Ai_POD-Storage"
    ]
    for col, value in enumerate(template_example, 1):
        template_sheet.cell(row=2, column=col, value=value)
        
    # Add organization dropdown to column B in Template sheet
    org_validation_template = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
    template_sheet.add_data_validation(org_validation_template)
    org_validation_template.add('B2:B1000')  # Column B
    
    # Add target platform dropdown to column E in Template sheet (correct column for Target Platform)
    platform_validation = DataValidation(type='list', formula1='"FIAttached,Standalone"', allow_blank=True)
    template_sheet.add_data_validation(platform_validation)
    platform_validation.add('E2:E1000')  # Column E (Target Platform)
    
    # Set up Profiles sheet - use the existing sheet
    profiles_sheet = workbook['Profiles']
    profile_headers = ["Profile Name*", "Description", "Organization*", "Resource Group*", "Template Name*", "Server*", "Description", "Deploy*"]
    for col, header in enumerate(profile_headers, 1):
        cell = profiles_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color='000000')  # Always black and bold
        cell.alignment = Alignment(horizontal='center')
    
    # Add 8 sample profile templates with Deploy set to No
    for i in range(1, 9):
        profiles_sheet.append([f'AI-Server-{i:02d}', 'AI POD Host Profile', 'default', 'AI POD Servers', 'Ai_POD_Template', '', f'Production AI POD Host {i}', 'No'])
    print(f"Added 8 profile templates to the Profiles sheet")
    
    # Add data validation for Deploy column
    deploy_validation = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
    profiles_sheet.add_data_validation(deploy_validation)
    deploy_validation.add('H2:H1000')  # Column H - Deploy column
    
    # Add organization dropdown to column C in Profiles sheet
    org_validation_profiles = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
    profiles_sheet.add_data_validation(org_validation_profiles)
    org_validation_profiles.add('C2:C1000')  # Column C
    
    # Add server dropdown (with name and serial) to column F
    server_validation = DataValidation(type='list', formula1='"Server-1 (FCH1234V5Z7),Server-2 (FCH5678A9BC),Server-3 (FCH9012D3EF)"', allow_blank=True)
    profiles_sheet.add_data_validation(server_validation)
    server_validation.add('F2:F1000')  # Column F - Server column
    
    # Create sample organization list for dropdowns
    org_list = ["default", "DevOps", "Production", "Test", "UAT"]
    
    # Create sample server list for dropdowns (with name and serial)
    server_list = ["Server-1 (FCH1234V5Z7)", "Server-2 (FCH5678A9BC)", "Server-3 (FCH9012D3EF)"]
    
    # Profiles sheet: Organization in column C, Server dropdown in column F, Deploy dropdown in column H
    org_validation_profiles = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
    profiles_sheet.add_data_validation(org_validation_profiles)
    org_validation_profiles.add('C2:C1000')  # Column C

    server_validation = DataValidation(type='list', formula1=f'"{",".join(server_list)}"', allow_blank=True)
    profiles_sheet.add_data_validation(server_validation)
    server_validation.add('F2:F1000')  # Column F only

    deploy_validation = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
    profiles_sheet.add_data_validation(deploy_validation)
    deploy_validation.add('H2:H1000')  # Column H only

    # Template name should not have a dropdown as it comes from Template sheet
    
    # Template sheet: Organization in column B
    org_validation_template = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
    template_sheet.add_data_validation(org_validation_template)
    org_validation_template.add('B2:B1000')  # Column B
    
    # Pools sheet: No organization dropdown needed (as per requirements)
    
    # Set column widths for all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # We're keeping only the essential sheets and dropdowns

    # Set up Servers sheet for server inventory
    servers_sheet = workbook['Servers']
    server_headers = ["Server Name*", "Serial Number*", "Description", "Model"]
    
    # Add headers to Servers sheet
    for col, header in enumerate(server_headers, 1):
        cell = servers_sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if '*' in header:
            cell.font = Font(color='000000', bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample server data
    sample_servers = [
        ("C220M5-Hosting-Server1", "FCH1234V5Z7", "Hosting Server 1", "UCS C220 M5"),
        ("C220M5-Hosting-Server2", "FCH5678A9BC", "Hosting Server 2", "UCS C220 M5"),
        ("C220M5-Hosting-Server3", "FCH9012D3EF", "Hosting Server 3", "UCS C220 M5")
    ]
    
    # Add servers to the sheet
    for row_idx, server in enumerate(sample_servers, 2):
        for col_idx, value in enumerate(server, 1):
            servers_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths for all sheets
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the workbook
    workbook.save(excel_file)
    print(f"Created template Excel file: {excel_file}")
    return True

def add_data_validation(worksheet, column, start_row, end_row, formula):
    """Helper function to add data validation to a worksheet"""
    validation = DataValidation(type='list', formula1=formula)
    worksheet.add_data_validation(validation)
    validation.add(f'{column}{start_row}:{column}{end_row}')

def add_dependency_sheet(workbook):
    """Add a policy dependency visualization sheet to the workbook.
    
    Args:
        workbook: The Excel workbook to add the dependency sheet to
    """
    try:
        # Check if Dependencies sheet exists
        if "Dependencies" not in workbook.sheetnames:
            # Create the sheet
            dep_sheet = workbook.create_sheet("Dependencies")
            
            # Add headers
            dep_sheet.cell(row=1, column=1, value="Policy/Component")
            dep_sheet.cell(row=1, column=2, value="Depends On")
            dep_sheet.cell(row=1, column=3, value="Relationship")
            
            # Format headers
            for col in range(1, 4):
                cell = dep_sheet.cell(row=1, column=col)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Add dependency data
            row = 2
            for component, dependencies in POLICY_DEPENDENCIES.items():
                for dependency in dependencies:
                    dep_sheet.cell(row=row, column=1, value=component)
                    dep_sheet.cell(row=row, column=2, value=dependency)
                    dep_sheet.cell(row=row, column=3, value="Required")
                    
                    # Apply formatting
                    for col in range(1, 4):
                        cell = dep_sheet.cell(row=row, column=col)
                        if col == 1:
                            cell.font = Font(bold=True)
                    
                    row += 1
            
            # Set column widths
            min_widths = {
                'A': 25,  # Policy Type
                'B': 25,  # Policy Name
                'C': 15   # Used By
            }
            # Use the auto-adjust function
            auto_adjust_column_width(dep_sheet, min_width=15, padding=2, custom_width_map=min_widths)
            
            print("Added Policy Dependencies visualization sheet")
            return True
        return False
    except Exception as e:
        print(f"Error adding dependency sheet: {str(e)}")
        return False

def add_version_sheet(workbook, version=TEMPLATE_VERSION):
    """Add or update a version sheet to track template changes.
    
    Args:
        workbook: The Excel workbook to add the version sheet to
        version: The version string to add to the sheet
    """
    try:
        # Check if Version sheet exists
        if "Version" not in workbook.sheetnames:
            # Create the sheet
            version_sheet = workbook.create_sheet("Version")
            # Add headers
            headers = ["Version", "Date", "Description", "Author"]
            for col, header in enumerate(headers, 1):
                cell = version_sheet.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Set column widths
            min_widths = {
                'A': 10,  # Version
                'B': 15,  # Date
                'C': 50,  # Description
                'D': 20   # Author
            }
            # Use the auto-adjust function
            auto_adjust_column_width(version_sheet, min_width=10, padding=2, custom_width_map=min_widths)
            
            # Add first entry
            version_sheet.cell(row=2, column=1, value=version)
            version_sheet.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d"))
            version_sheet.cell(row=2, column=3, value="Initial template creation with dynamic organization and server dropdowns")
            version_sheet.cell(row=2, column=4, value=os.environ.get('USER', 'Intersight-Admin'))
        else:
            # Update existing version sheet with new entry
            version_sheet = workbook["Version"]
            next_row = version_sheet.max_row + 1
            version_sheet.cell(row=next_row, column=1, value=version)
            version_sheet.cell(row=next_row, column=2, value=datetime.now().strftime("%Y-%m-%d"))
            version_sheet.cell(row=next_row, column=3, value="Updated template with latest organizations and servers")
            version_sheet.cell(row=next_row, column=4, value=os.environ.get('USER', 'Intersight-Admin'))
        
        return True
    except Exception as e:
        print(f"Error adding version sheet: {str(e)}")
        return False

def get_intersight_info(api_client, excel_file):
    """Get information from Intersight and update the Excel file"""
    try:
        # Load existing workbook
        workbook = load_workbook(excel_file)
        
        # Correct sheet naming and order
        sheet_renames = {
            'Pools': 'Pools1',
            'Policies': 'Policies1',
            'Template': 'Template1',
            'Profiles': 'Profiles1',
            'Pools1': 'Pools',
            'Policies1': 'Policies',
            'Template1': 'Template',
            'Profiles1': 'Profiles'
        }
        
        for old_name, new_name in sheet_renames.items():
            if old_name in workbook.sheetnames:
                workbook[old_name].title = new_name
        
        # DO NOT rearrange sheet order to preserve template structure
        # Just log the current sheets
        print(f"Working with existing sheets: {', '.join(workbook.sheetnames)}")

        # Get organizations
        print("\nGetting organizations from Intersight...")
        org_api = organization_api.OrganizationApi(api_client)
        orgs = org_api.get_organization_organization_list()
        org_names = [org.name for org in orgs.results]
        print(f"Found {len(org_names)} organizations: {org_names}")

        # Get resource groups
        print("\nGetting resource groups from Intersight...")
        try:
            resource_api_instance = resource_api.ResourceApi(api_client)
            resource_groups = resource_api_instance.get_resource_group_list()
            resource_group_names = [group.name for group in resource_groups.results]
            if not resource_group_names:
                resource_group_names = ["Default", "AI POD Servers", "ML Servers", "DevOps", "Production"]
                print(f"No resource groups found. Using default sample values: {resource_group_names}")
            else:
                print(f"Found {len(resource_group_names)} resource groups: {resource_group_names}")
        except Exception as e:
            resource_group_names = ["Default", "AI POD Servers", "ML Servers", "DevOps", "Production"]
            print(f"Error getting resource groups: {str(e)}. Using default sample values: {resource_group_names}")

        # Get servers
        print("\nGetting servers from Intersight...")
        compute_api_instance = compute_api.ComputeApi(api_client)
        servers = compute_api_instance.get_compute_rack_unit_list()
        server_names = [server.name for server in servers.results]
        print(f"Found {len(server_names)} servers: {server_names}")
        
        # Populate Servers sheet
        if 'Servers' in workbook.sheetnames:
            servers_sheet = workbook['Servers']
            
            # Find the header row
            header_row = 1
            for row in range(1, min(5, servers_sheet.max_row + 1)):
                if "Server Name" in str(servers_sheet.cell(row=row, column=1).value):
                    header_row = row
                    break
            
            # Clear existing data but only below the header row
            for row in range(header_row + 1, servers_sheet.max_row + 1):
                for col in range(1, 5):
                    servers_sheet.cell(row=row, column=col).value = None
            
            # Add actual server data starting right after the header row
            row = header_row + 1
            for i, server_name in enumerate(server_names):
                servers_sheet.cell(row=row, column=1, value=server_name)
                try:
                    server_filter = f"Name eq '{server_name}'"
                    server_details = compute_api_instance.get_compute_rack_unit_list(filter=server_filter)
                    if server_details.results and len(server_details.results) > 0:
                        server = server_details.results[0]
                        servers_sheet.cell(row=row, column=2, value=server.serial)
                        servers_sheet.cell(row=row, column=3, value=f"Intersight managed server")
                        servers_sheet.cell(row=row, column=4, value=server.model)
                    else:
                        servers_sheet.cell(row=row, column=2, value=f"FCH{(i+1)*12345:07d}")
                        servers_sheet.cell(row=row, column=3, value=f"Intersight managed server {i+1}")
                        model = "UCS C-Series"
                        if "C220M5" in server_name:
                            model = "UCS C220 M5"
                        elif "C220M4" in server_name:
                            model = "UCS C220 M4"
                        elif "C480ML" in server_name:
                            model = "UCS C480 ML M5"
                        servers_sheet.cell(row=row, column=4, value=model)
                except Exception as e:
                    print(f"  - Warning: Could not get detailed info for server {server_name}: {str(e)}")
                    servers_sheet.cell(row=row, column=2, value=f"FCH{(i+1)*12345:07d}")
                    servers_sheet.cell(row=row, column=3, value=f"Intersight managed server {i+1}")
                    servers_sheet.cell(row=row, column=4, value="UCS C-Series")
                row += 1
        
            # Reapply header formatting
            from openpyxl.styles import PatternFill, Font, Alignment
            header_fill = PatternFill(start_color='A0D7BE', end_color='A0D7BE', fill_type='solid')
            for col in range(1, 5):
                cell = servers_sheet.cell(row=header_row, column=col)
                cell.font = Font(bold=True, color='FF000000')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # Set up Profiles sheet dropdowns
        if 'Profiles' in workbook.sheetnames:
            profiles_sheet = workbook['Profiles']
            
            # Ensure header row formatting is correct
            # Define styles - using a lighter shade of green for a more subtle look
            header_fill = PatternFill(start_color='A0D7BE', end_color='A0D7BE', fill_type='solid')  # Light green
            header_font = Font(color='FF000000', bold=True)  # Black text for readability
            
            # Re-apply formatting to headers
            for col in range(1, profiles_sheet.max_column + 1):
                cell = profiles_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
            
            # Clear all validations
            profiles_sheet.data_validations.dataValidation = []
            
            # Re-add existing non-org validations
            existing_dv = []
            org_dv_found = False
            
            for dv in list(profiles_sheet.data_validations.dataValidation):
                # Check if this is an organization dropdown
                is_org_dv = False
                for cell_range in dv.sqref:
                    if str(cell_range).startswith('C'):
                        # This is the org dropdown, update it
                        dv.formula1 = f'"{",".join(org_names)}"'
                        org_dv_found = True
                        is_org_dv = True
                        break
                
                # Keep all non-org dropdowns
                if not is_org_dv:
                    existing_dv.append(dv)
            
            # Clear and re-add all validations
            profiles_sheet.data_validations.dataValidation = []
            
            # Re-add existing non-org validations
            for dv in existing_dv:
                profiles_sheet.add_data_validation(dv)
            
            # Find column indexes by header name
            header_map = {profiles_sheet.cell(row=1, column=col).value: col for col in range(1, profiles_sheet.max_column + 1)}
            server_col = header_map.get('Server*')
            deploy_col = header_map.get('Deploy*')
            org_col = header_map.get('Organization*')
            resource_group_col = header_map.get('Resource Group*')
            template_name_col = header_map.get('Template Name*')
            
            # Add server dropdown to correct column
            if server_col:
                server_options = [f"{server.name} | SN: {server.serial}" for server in servers.results]
                server_formula = '"' + ','.join(server_options) + '"'
                server_validation = DataValidation(
                    type='list',
                    formula1=server_formula,
                    allow_blank=True
                )
                col_letter = get_column_letter(server_col)
                server_validation.add(f'{col_letter}2:{col_letter}1000')
                profiles_sheet.add_data_validation(server_validation)
            
            # Add deploy dropdown to correct column
            if deploy_col:
                deploy_validation = DataValidation(
                    type='list',
                    formula1='"Yes,No"',
                    allow_blank=True
                )
                col_letter = get_column_letter(deploy_col)
                deploy_validation.add(f'{col_letter}2:{col_letter}1000')
                profiles_sheet.add_data_validation(deploy_validation)
            
            # Add organization dropdown to correct column
            if org_col:
                print(f"Updating organization dropdown options with values: {org_names}")
                org_formula = f'"{",".join(org_names)}"'
                print(f"Organization formula: {org_formula}")
                org_validation = DataValidation(
                    type='list',
                    formula1=org_formula,
                    allow_blank=True
                )
                col_letter = get_column_letter(org_col)
                org_validation.add(f'{col_letter}2:{col_letter}1000')
                profiles_sheet.add_data_validation(org_validation)
            
            # Add resource group dropdown to correct column
            if resource_group_col:
                resource_group_formula = '"' + ','.join(resource_group_names) + '"'
                resource_group_validation = DataValidation(
                    type='list',
                    formula1=resource_group_formula,
                    allow_blank=True
                )
                col_letter = get_column_letter(resource_group_col)
                resource_group_validation.add(f'{col_letter}2:{col_letter}1000')
                profiles_sheet.add_data_validation(resource_group_validation)
            
            # Add template name dropdown to correct column
            if template_name_col:
                # Gather template names from the Template sheet if available
                template_names = []
                if 'Template' in workbook.sheetnames:
                    template_sheet = workbook['Template']
                    for row in range(2, template_sheet.max_row + 1):
                        name = template_sheet.cell(row=row, column=1).value
                        if name:
                            template_names.append(str(name))
                if template_names:
                    template_name_formula = '"' + ','.join(template_names) + '"'
                    template_name_validation = DataValidation(
                        type='list',
                        formula1=template_name_formula,
                        allow_blank=True
                    )
                    col_letter = get_column_letter(template_name_col)
                    template_name_validation.add(f'{col_letter}2:{col_letter}1000')
                    profiles_sheet.add_data_validation(template_name_validation)
            
            print("Added/Updated dropdowns for Server, Deploy, Organization, Resource Group, and Template Name columns")
        
        # Set up dropdowns for all sheets
        # Pools sheet dropdown
        if 'Pools' in workbook.sheetnames:
            pools_sheet = workbook['Pools']
            pool_types = ['MAC Pool', 'UUID Pool']
            pool_validation = DataValidation(
                type='list',
                formula1=f'"{",".join(pool_types)}"',
                allow_blank=True
            )
            pool_validation.add('A2:A1000')  # Apply to Pool Types column
            pools_sheet.add_data_validation(pool_validation)
            print("Added dropdown for Pool Types in Pools sheet")

        # Policies sheet dropdown
        if 'Policies' in workbook.sheetnames:
            policies_sheet = workbook['Policies']
            
            # Always create fresh organization dropdown
            print(f"Updating organization dropdown for Policies sheet with values: {org_names}")
            org_formula = f'"{",".join(org_names)}"'
            org_validation = DataValidation(
                type='list',
                formula1=org_formula,
                allow_blank=True
            )
            org_validation.add('D2:D1000')  # Apply to Organizations columns
            policies_sheet.add_data_validation(org_validation)
            
            print("Added/Updated dropdowns for Policy Types and Organizations in Policies sheet")
        
        # Template sheet dropdowns
        if 'Template' in workbook.sheetnames:
            template_sheet = workbook['Template']
            
            # Always create fresh organization dropdown
            print(f"Updating organization dropdown for Template sheet with values: {org_names}")
            org_formula = f'"{",".join(org_names)}"'
            org_validation = DataValidation(
                type='list',
                formula1=org_formula,
                allow_blank=True
            )
            org_validation.add('B2:B1000')  # Apply to Organizations column
            template_sheet.add_data_validation(org_validation)
            
            print("Added/Updated dropdowns for Platform Types and Organizations in Template sheet")
        
        # Skip auto-adjusting column widths to preserve template formatting
        print("\nPreserving column widths to maintain template formatting...")
        
        # Save workbook
        print("\nSaving Excel file...")
        workbook.save(excel_file)
        print("Excel file has been set up with correct sheet order and structure")
        return True
    except Exception as e:
        print(f"Error setting up Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def get_org_moid(api_client, org_name="Gruve"):  # Set default to Gruve
    """
    Get the MOID (Managed Object ID) for an organization by name
    """
    from intersight.api import organization_api
    
    try:
        # Create Organization API instance
        api_instance = organization_api.OrganizationApi(api_client)
        
        # Get list of organizations
        orgs = api_instance.get_organization_organization_list(filter=f"Name eq '{org_name}'")
        
        if orgs.results and len(orgs.results) > 0:
            return orgs.results[0].moid
        else:
            raise Exception(f"Organization '{org_name}' not found")
            
    except Exception as e:
        raise Exception(f"Error getting organization MOID: {str(e)}")

def create_policy(api_client, policy_data):
    """
    Create a policy in Intersight based on the provided data
    """
    policy_type = policy_data['Policy Type']
    policy_name = policy_data['Policy Name']  # Updated from 'Name' to 'Policy Name'
    
    try:
        # Get Gruve organization MOID
        org_moid = get_org_moid(api_client, "Gruve")
        if not org_moid:
            print("Error: Gruve organization not found")
            return False

        org_ref = {
            "class_id": "mo.MoRef",
            "object_type": "organization.Organization",
            "moid": org_moid
        }
        
        print(f"\nCreating {policy_type} policy: {policy_name}")
        
        if policy_type == 'BIOS':
            from intersight.api import bios_api
            from intersight.model.bios_policy import BiosPolicy
            
            api_instance = bios_api.BiosApi(api_client)
            
            # Create BIOS policy with performance settings
            policy = BiosPolicy(
                class_id="bios.Policy",
                object_type="bios.Policy",
                name=policy_name,
                organization=org_ref,
                cpu_performance="enterprise",
                cpu_power_management="performance",
                cpu_energy_performance="performance",
                intel_virtualization_technology="enabled"
            )
            
            result = api_instance.create_bios_policy(policy)
            print(f"Successfully created BIOS Policy: {result.name}")
            return True
            
        elif policy_type == 'QoS':
            from intersight.api import vnic_api
            
            api_instance = vnic_api.VnicApi(api_client)
            
            # Create QoS policy
            qos = {
                "name": policy_name,
                "organization": org_ref,
                "mtu": 9000,
                "rate_limit": 0,
                "cos": 5,
                "burst": 1024,
                "priority": "Best Effort",
                "class_id": "vnic.EthQosPolicy",
                "object_type": "vnic.EthQosPolicy"
            }
            
            result = api_instance.create_vnic_eth_qos_policy(qos)
            print(f"Successfully created QoS Policy: {result.name}")
            return True
            
        elif policy_type == 'vNIC':
            from intersight.api import vnic_api, fabric_api
            from intersight.model.vnic_lan_connectivity_policy import VnicLanConnectivityPolicy
            from intersight.model.vnic_eth_if import VnicEthIf
            
            # Create API instances
            vnic_instance = vnic_api.VnicApi(api_client)
            fabric_instance = fabric_api.FabricApi(api_client)
            
            # Create Ethernet Adapter Policy
            eth_adapter = {
                "class_id": "vnic.EthAdapterPolicy",
                "object_type": "vnic.EthAdapterPolicy",
                "name": f"{policy_name}-eth-adapter",
                "organization": org_ref,
                "rss_settings": True,
                "uplink_failback_timeout": 5,
                "interrupt_settings": {
                    "coalescing_time": 125,
                    "coalescing_type": "MIN",
                    "count": 4,
                    "mode": "MSIx"
                },
                "rx_queue_settings": {
                    "count": 1,
                    "ring_size": 512
                },
                "tx_queue_settings": {
                    "count": 1,
                    "ring_size": 256
                },
                "completion_queue_settings": {
                    "count": 2,
                    "ring_size": 1
                },
                "tcp_offload_settings": {
                    "large_receive": True,
                    "large_send": True,
                    "rx_checksum": True,
                    "tx_checksum": True
                },
                "advanced_filter": True
            }
            
            eth_adapter_result = vnic_instance.create_vnic_eth_adapter_policy(eth_adapter)
            print(f"Successfully created Ethernet Adapter Policy: {eth_adapter_result.name}")

            # Create Network Group Policies for Fabric A and B
            network_group_a = {
                "class_id": "fabric.EthNetworkGroupPolicy",
                "object_type": "fabric.EthNetworkGroupPolicy",
                "name": f"{policy_name}-network-group-A",
                "organization": org_ref,
                "vlan_settings": {
                    "native_vlan": 1,
                    "allowed_vlans": "2-100"
                }
            }
            
            network_group_b = {
                "class_id": "fabric.EthNetworkGroupPolicy",
                "object_type": "fabric.EthNetworkGroupPolicy",
                "name": f"{policy_name}-network-group-B",
                "organization": org_ref,
                "vlan_settings": {
                    "native_vlan": 1,
                    "allowed_vlans": "2-100"
                }
            }
            
            group_a_result = fabric_instance.create_fabric_eth_network_group_policy(network_group_a)
            print(f"Successfully created Network Group Policy A: {group_a_result.name}")
            
            group_b_result = fabric_instance.create_fabric_eth_network_group_policy(network_group_b)
            print(f"Successfully created Network Group Policy B: {group_b_result.name}")

            # Create vNIC Policy
            lan_connectivity = {
                "class_id": "vnic.LanConnectivityPolicy",
                "object_type": "vnic.LanConnectivityPolicy",
                "name": policy_name,
                "organization": org_ref,
                "target_platform": "FIAttached"
            }
            
            lan_policy = vnic_instance.create_vnic_lan_connectivity_policy(lan_connectivity)
            print(f"Successfully created vNIC LAN Connectivity Policy: {lan_policy.name}")

            # Create vNIC eth0 for Fabric A
            eth0 = {
                "class_id": "vnic.EthIf",
                "object_type": "vnic.EthIf",
                "name": f"eth0_{lan_policy.name}",  # Make the name unique
                "order": 0,
                "placement": {
                    "class_id": "vnic.PlacementSettings",
                    "object_type": "vnic.PlacementSettings",
                    "id": "MLOM",
                    "pci_link": 0,
                    "switch_id": "A",
                    "uplink": 0
                },
                "cdn": {
                    "class_id": "vnic.Cdn",
                    "object_type": "vnic.Cdn",
                    "source": "vnic",
                    "value": "eth0"
                },
                "eth_qos_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.EthQosPolicy",
                    "moid": get_policy_moid(api_client, "vnic.EthQosPolicy", "Ai_POD-QoS")
                },
                "eth_adapter_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.EthAdapterPolicy",
                    "moid": eth_adapter_result.moid
                },
                "fabric_eth_network_group_policy": [{
                    "class_id": "mo.MoRef",
                    "object_type": "fabric.EthNetworkGroupPolicy",
                    "moid": group_a_result.moid
                }],
                "lan_connectivity_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.LanConnectivityPolicy",
                    "moid": lan_policy.moid
                },
                "mac_pool": {
                    "class_id": "mo.MoRef",
                    "object_type": "macpool.Pool",
                    "moid": get_mac_pool_moid(api_client, "Ai_POD-MAC-A", org_moid)
                }
            }

            # Create vNIC eth1 for Fabric B
            eth1 = {
                "class_id": "vnic.EthIf",
                "object_type": "vnic.EthIf",
                "name": f"eth1_{lan_policy.name}",  # Make the name unique
                "order": 1,
                "placement": {
                    "class_id": "vnic.PlacementSettings",
                    "object_type": "vnic.PlacementSettings",
                    "id": "MLOM",
                    "pci_link": 1,  # Different PCI link for eth1
                    "switch_id": "B",
                    "uplink": 0
                },
                "cdn": {
                    "class_id": "vnic.Cdn",
                    "object_type": "vnic.Cdn",
                    "source": "vnic",
                    "value": "eth1"
                },
                "eth_qos_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.EthQosPolicy",
                    "moid": get_policy_moid(api_client, "vnic.EthQosPolicy", "Ai_POD-QoS")
                },
                "eth_adapter_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.EthAdapterPolicy",
                    "moid": eth_adapter_result.moid
                },
                "fabric_eth_network_group_policy": [{
                    "class_id": "mo.MoRef",
                    "object_type": "fabric.EthNetworkGroupPolicy",
                    "moid": group_b_result.moid
                }],
                "lan_connectivity_policy": {
                    "class_id": "mo.MoRef",
                    "object_type": "vnic.LanConnectivityPolicy",
                    "moid": lan_policy.moid
                },
                "mac_pool": {
                    "class_id": "mo.MoRef",
                    "object_type": "macpool.Pool",
                    "moid": get_mac_pool_moid(api_client, "Ai_POD-MAC-B", org_moid)
                }
            }

            # Create the vNICs
            eth0_name = f"eth0_{lan_policy.name}"
            eth1_name = f"eth1_{lan_policy.name}"

            if not check_vnic_exists(api_client, eth0_name, lan_policy.moid):
                print("\nCreating vNIC eth0 for Fabric A...")
                eth0_result = vnic_instance.create_vnic_eth_if(eth0)
                print(f"Successfully created vNIC eth0 for Fabric A")
            else:
                print(f"\nvNIC {eth0_name} already exists, skipping creation")

            if not check_vnic_exists(api_client, eth1_name, lan_policy.moid):
                print("\nCreating vNIC eth1 for Fabric B...")
                eth1_result = vnic_instance.create_vnic_eth_if(eth1)
                print(f"Successfully created vNIC eth1 for Fabric B")
            else:
                print(f"\nvNIC {eth1_name} already exists, skipping creation")
            
            return True
            
        elif policy_type == 'Storage':
            from intersight.api import storage_api
            from intersight.model.storage_storage_policy import StorageStoragePolicy
            from intersight.model.storage_virtual_drive_policy import StorageVirtualDrivePolicy
            from intersight.model.storage_r0_drive import StorageR0Drive
            
            api_instance = storage_api.StorageApi(api_client)
            
            # Create virtual drive policy
            virtual_drive_policy = StorageVirtualDrivePolicy(
                drive_cache="Default",
                read_policy="Default",
                strip_size=512,
                access_policy="Default"
            )
            
            # Create RAID0 drive configuration
            raid0_drive = StorageR0Drive(
                enable=True,
                virtual_drive_policy=virtual_drive_policy
            )
            
            # Create storage policy
            storage_pol = StorageStoragePolicy(
                name=policy_name,
                description=policy_data.get('Description', ''),
                organization=org_ref,
                default_drive_mode="RAID0",
                raid0_drive=raid0_drive,
                unused_disks_state="NoChange",
                use_jbod_for_vd_creation=False
            )
            
            try:
                result = api_instance.create_storage_storage_policy(storage_storage_policy=storage_pol)
                print(f"Successfully created Storage Policy: {result.name}")
                return True
            except Exception as e:
                print(f"Error creating Storage policy: {str(e)}")
                raise
            
        elif policy_type == 'Boot':
            from intersight.api import boot_api
            from intersight.model.boot_precision_policy import BootPrecisionPolicy
            from intersight.model.boot_device_base import BootDeviceBase
            from intersight.model.boot_uefi_shell import BootUefiShell
            from intersight.model.boot_pxe import BootPxe
            
            api_instance = boot_api.BootApi(api_client)
            
            # Create UEFI Shell boot device
            boot_uefi = BootUefiShell(
                class_id="boot.UefiShell",
                object_type="boot.UefiShell",
                name="uefi1",
                enabled=True
            )
            
            # Create PXE boot device
            boot_pxe = BootPxe(
                class_id="boot.Pxe",
                object_type="boot.Pxe",
                name="pxe1",
                interface_name="eth0",
                ip_type="IPv4",
                enabled=True
            )
            
            # Create local disk boot device
            boot_local_disk = BootDeviceBase(
                class_id="boot.LocalDisk",
                object_type="boot.LocalDisk",
                name="local_disk1",
                enabled=True
            )
            
            # Create boot devices list
            boot_devices = [
                boot_local_disk,
                boot_uefi,
                boot_pxe
            ]
            
            # Create boot policy with the boot devices
            boot_pol = BootPrecisionPolicy(
                name=policy_name,
                description=policy_data.get('Description', ''),
                organization=org_ref,
                configured_boot_mode="Uefi",
                boot_devices=boot_devices
            )
            
            try:
                result = api_instance.create_boot_precision_policy(boot_precision_policy=boot_pol)
                print(f"Successfully created Boot Policy: {result.name}")
                return True
            except Exception as e:
                print(f"Error creating Boot policy: {str(e)}")
                raise
            
        else:
            print(f"Unsupported policy type: {policy_type}")
            return False
            
    except Exception as e:
        print(f"Error creating {policy_type} policy: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def update_profiles_with_server_info(api_client, excel_file):
    """Update the Profiles sheet with server information from Intersight"""
    try:
        # Load workbook
        workbook = load_workbook(excel_file)
        if 'Profiles' not in workbook.sheetnames:
            print("No Profiles sheet found in Excel file")
            return False
        
        profiles_sheet = workbook['Profiles']
        
        # Get servers from Intersight
        compute_api_instance = compute_api.ComputeApi(api_client)
        servers = compute_api_instance.get_compute_rack_unit_list()
        
        # Collect server info for dropdown
        server_options = [f"{server.name} | SN: {server.serial}" for server in servers.results]
        server_list = ','.join(server_options)
        
        # Add server dropdown to row 2
        server_validation = DataValidation(
            type='list',
            formula1=f'"{server_list}"',
            allow_blank=True
        )
        profiles_sheet.add_data_validation(server_validation)
        server_validation.add('E2')
        print("Added server dropdown to row 2 in Profiles sheet")
        
        # Save workbook
        try:
            workbook.save(excel_file)
            print("Successfully saved Excel file")
        except Exception as e:
            print(f"Failed to save Excel file: {str(e)}")
        
        return True
        
    except Exception as e:
        print(f"Error updating Profiles sheet: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# Define retry decorator directly in script to avoid import issues
def retry_api_call(max_retries=3, delay=2):
    """Decorator to retry API calls with exponential backoff"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            current_delay = delay
            while retries < max_retries:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    retries += 1
                    if retries >= max_retries:
                        print(f"API call failed after {max_retries} attempts: {str(e)}")
                        raise
                    
                    print(f"API call failed. Retrying in {current_delay}s... ({retries}/{max_retries})")
                    time.sleep(current_delay)
                    current_delay *= 1.5  # Exponential backoff
        return wrapper
    return decorator

# Fallback print functions if utils module import failed
def print_info(message):
    print(message)
    
def print_success(message):
    print(message)
    
def print_warning(message):
    print(message)
    
def print_error(message):
    print(message)

def progress_bar(iterable, desc="", total=None):
    return iterable

def print_summary(title, success_items, failed_items):
    print(f"\n{title} Summary")
    if success_items:
        print(f"Successfully processed {len(success_items)} items")
    if failed_items:
        print(f"Failed to process {len(failed_items)} items")
        
def validate_pools_data(pools_df):
    """Validate pools data before creating in Intersight"""
    invalid_pools = []
    
    # Debug: Print column names to ensure we're looking for the right columns
    print("DEBUG: Pool columns available:", pools_df.columns.tolist())
    print("\nDEBUG: First 5 rows of pool data:")
    for i, (idx, row) in enumerate(pools_df.iterrows()):
        if i < 5:  # Only show first 5 rows
            print(f"DEBUG: Row {idx+2} data: {dict(row)}")
    
    for idx, row in pools_df.iterrows():
        pool_type = row.get('Pool Type')
        pool_name = row.get('Pool Name')
        
        # Debug: Print the values being processed
        print(f"\nValidating Row {idx+2}:")
        print(f"  Pool Type: '{pool_type}'")
        print(f"  Pool Name: '{pool_name}'")
        
        # Check for missing required fields
        if not pool_type or pd.isna(pool_type):
            invalid_pools.append(f"Row {idx+2}: Missing Pool Type")
            continue
        
        if not pool_name or pd.isna(pool_name):
            invalid_pools.append(f"Row {idx+2}: Missing Pool Name")
            continue
            
        # Validate pool type specific fields
        if pool_type == 'MAC Pool':
            # Check for either 'Start Address' or 'First Address' as the column name
            start_addr = row.get('Start Address')
            if start_addr is None or pd.isna(start_addr):
                start_addr = row.get('First Address')  # Try the alternate column name
            
            size = row.get('Size')
            
            # Debug: Print the MAC Pool specific values
            print(f"  Start/First Address: '{start_addr}' (type: {type(start_addr).__name__})")
            print(f"  Size: '{size}' (type: {type(size).__name__})")
            
            if not start_addr or pd.isna(start_addr):
                invalid_pools.append(f"Row {idx+2}: Missing Start/First Address for MAC Pool '{pool_name}'")
            elif not isinstance(start_addr, str):
                invalid_pools.append(f"Row {idx+2}: Invalid Start/First Address format for MAC Pool '{pool_name}'")
            
            if not size or pd.isna(size):
                invalid_pools.append(f"Row {idx+2}: Missing Size for MAC Pool '{pool_name}'")
            elif not str(size).isdigit():
                invalid_pools.append(f"Row {idx+2}: Size must be a number for MAC Pool '{pool_name}'")
    
    return invalid_pools

def validate_policies_data(policies_df):
    """Validate policies data before creating in Intersight"""
    invalid_policies = []
    
    for idx, row in policies_df.iterrows():
        policy_type = row.get('Policy Type')
        policy_name = row.get('Policy Name')
        
        # Check for missing required fields
        if not policy_type or pd.isna(policy_type):
            invalid_policies.append(f"Row {idx+2}: Missing Policy Type")
            continue
        
        if not policy_name or pd.isna(policy_name):
            invalid_policies.append(f"Row {idx+2}: Missing Policy Name")
            continue
    
    return invalid_policies

@retry_api_call(max_retries=3, delay=2)
def create_and_derive_profile(api_client, profile_data):
    """Create a server profile and then attach it to a template using the official API approach"""
    from intersight.api import server_api
    from intersight.model.server_profile import ServerProfile
    from intersight.model.mo_mo_ref import MoMoRef
    
    # Map DataFrame column names to expected parameter names
    profile_name = profile_data.get('Profile Name')
    template_name = profile_data.get('Template Name')
    server_name = profile_data.get('Server')
    org_name = profile_data.get('Organization')
    deploy = profile_data.get('Deploy', 'No')
    description = f"Server Profile for {server_name}"
    
    print_info(f"\nCreating server profile: {profile_name}")
    if deploy.lower() == "yes":
        print_info(f"Profile {profile_name} will be deployed after creation")
    
    print_info(f"Organization: {org_name}")
    print_info(f"Template: {template_name}")
    print_info(f"Server: {server_name}")
    
    try:
        # Create API instance
        api_instance = server_api.ServerApi(api_client)
        
        # Get organization MOID
        print_info(f"Looking up organization: {org_name}")
        org_moid = get_org_moid(api_client, org_name)
        if not org_moid:
            print_error(f"Organization {org_name} not found")
            return False
            
        # Create organization reference
        organization = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        print_success(f"Found organization with MOID: {org_moid}")
        
        # Get template MOID - check if we have a template mapping for this name
        if template_mappings and template_name in template_mappings:
            mapped_template_name = template_mappings[template_name]
            print(f"Found template mapping for {template_name} -> {mapped_template_name}")
            template_name = mapped_template_name
        
        template_moid = get_template_moid(api_client, template_name)
        if not template_moid:
            print(f"Template {template_name} not found")
            return False
        
        # Create template reference
        template_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="server.ProfileTemplate",
            moid=template_moid
        )
        
        # Get server MOID if specified
        server_ref = None
        if server_name:
            # Extract serial number if format is "Name | SN: XYZ"
            serial_number = None
            if " | SN: " in server_name:
                parts = server_name.split(" | SN: ")
                server_name = parts[0].strip()
                serial_number = parts[1].strip()
            
            server_moid = get_server_moid(api_client, server_name)
            if not server_moid:
                print(f"Server {server_name} not found")
                return False
                
            # Create server reference
            server_ref = MoMoRef(
                class_id="mo.MoRef",
                object_type="compute.RackUnit",
                moid=server_moid
            )
        
        # STEP 1: Create ServerProfile instance following the official docs
        print("Creating server profile using official API approach...")
        server_profile = ServerProfile()
        server_profile.name = profile_name
        server_profile.description = description
        server_profile.organization = organization
        server_profile.target_platform = "Standalone"  # Assuming standalone for now
        server_profile.type = "instance"  # 'instance' for profiles
        
        # Don't add server during profile creation - we'll do it after template attachment
        
        # Create the profile
        print(f"Creating profile: {profile_name}")
        resp_server_profile = api_instance.create_server_profile(server_profile)
        profile_moid = resp_server_profile.moid
        print(f"Successfully created profile with MOID: {profile_moid}")
        
        # STEP 2: Update profile to attach it to the template
        print(f"Attaching profile to template {template_name}...")
        
        # Create update body with template reference
        update_profile = ServerProfile()
        update_profile.src_template = template_ref
        
        # Update the profile to attach to template
        api_instance.update_server_profile(profile_moid, update_profile)
        
        print(f"Successfully created and attached profile {profile_name} to template")
        return True
        
    except Exception as e:
        print(f"Error using official API approach: {str(e)}")
        print("\n⚠️ Unable to create server profile with template attachment.")
        print("This profile will need to be created manually in the Intersight UI.")
        
        # Store profile for manual creation report
        if 'profiles_for_manual_creation' not in globals():
            global profiles_for_manual_creation
            profiles_for_manual_creation = []
        
        # Store info needed for manual creation
        profile_info = {
            'name': profile_name,
            'template': template_name,
            'organization': org_name,
            'server': server_name,
            'deploy': deploy
        }
        profiles_for_manual_creation.append(profile_info)
        
        return False

def create_basic_server_profile(api_client, profile_name, org_moid, server_moid=None):
    """Create a basic server profile"""
    from intersight.api import server_api
    from intersight.model.server_profile import ServerProfile
    from intersight.model.mo_mo_ref import MoMoRef
    
    try:
        # Create organization reference
        org_ref = MoMoRef(
            class_id="mo.MoRef",
            object_type="organization.Organization",
            moid=org_moid
        )
        
        # Create server reference if server_moid exists
        server_ref = None
        if server_moid:
            server_ref = MoMoRef(
                class_id="mo.MoRef",
                object_type="compute.RackUnit",
                moid=server_moid
            )
        
        # Create API instance
        api_instance = server_api.ServerApi(api_client)
        
        # Create the profile using the SDK model objects
        server_profile = ServerProfile()
        server_profile.name = str(profile_name)
        server_profile.type = "instance"
        
        # Add organization reference with proper model object
        server_profile.organization = org_ref
        # Add server if specified
        if server_moid:
            profile_min["AssignedServer"] = {
                "ClassId": "mo.MoRef",
                "ObjectType": "compute.RackUnit",
                "Moid": server_moid
            }
        
        # Configure HTTP request args
        local_var_params = {'server_profile': profile_min}
        collection_formats = {}
        path_params = {}
        query_params = []
        header_params = {}
        form_params = []
        local_var_files = {}
        body_params = None
        
        # HTTP header `Accept`
        header_params['Accept'] = api_client.select_header_accept(['application/json'])
        # HTTP header `Content-Type`
        header_params['Content-Type'] = api_client.select_header_content_type(['application/json'])
        
        # Authentication setting
        auth_settings = ['cookieAuth', 'oAuth2']
        body_params = local_var_params['server_profile']
        
        # Make direct call to API client
        api_client.call_api(
            '/server/Profiles', 'POST',
            path_params,
            query_params,
            header_params,
            body=body_params,
            post_params=form_params,
            files=local_var_files,
            response_type='ServerProfile',
            auth_settings=auth_settings,
            async_req=False,
            _return_http_data_only=True,
            _preload_content=True,
            collection_formats=collection_formats)
        
        print(f"Successfully created Server Profile: {profile_name}")
        return profile_name
    
    except Exception as e:
        print(f"Error creating Server Profile: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def derive_profile_from_template(api_client, profile_moid, template_moid):
    """Derive a server profile from a template"""
    from intersight.api import server_api
    
    try:
        # Create API instance
        api_instance = server_api.ServerApi(api_client)
        
        # Derive the profile from the template
        api_instance.derive_server_profile(server_profile_moid=profile_moid, server_profile_template_moid=template_moid)
        print(f"Successfully derived Server Profile from Template")
        return True
    
    except Exception as e:
        print(f"Error deriving Server Profile from Template: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Create and push Intersight Foundation configuration')
    parser.add_argument('--action', choices=['push', 'template', 'profiles', 'all', 'setup', 'create-template', 'get-info', 'update-servers'], required=True,
                      help='Action to perform: push (create pools and policies), template (create server template), profiles (create server profiles), all (do everything), setup (just set up Excel file), create-template (create fresh template), get-info (get current Intersight information), update-servers (update server info in Profiles sheet)')
    parser.add_argument('--file', default='output/Intersight_Template.xlsx', help='Path to Excel file (default: output/Intersight_Template.xlsx)')
    args = parser.parse_args()
    
    if args.action == 'update-servers':
        api_client = get_api_client()
        if not api_client:
            sys.exit(1)
        update_profiles_with_server_info(api_client, args.file)
    elif args.action == 'create-template':
        # Always use Intersight_Template.xlsx as the standard filename unless specified otherwise
        if args.file != 'output/Intersight_Template.xlsx' and '--file' not in sys.argv:
            print(f"Using standard template filename: output/Intersight_Template.xlsx")
            args.file = 'output/Intersight_Template.xlsx'
            
        # Ensure output directory exists
        os.makedirs(os.path.dirname(args.file), exist_ok=True)
            
        if os.path.exists(args.file):
            response = input(f"A file named '{args.file}' already exists. Would you like to keep the existing file or replace it? (k = keep, r = replace): ").strip().lower()
            if response == 'k':
                print(f"Keeping the existing file: {args.file}. No changes made.")
                sys.exit(0)
            elif response == 'r':
                print(f"Replacing the existing file: {args.file}.")
                create_template_excel(args.file)
            else:
                print("Invalid response. No changes made.")
                sys.exit(1)
        else:
            create_template_excel(args.file)
    elif args.action == 'setup':
        api_client = get_api_client()
        if not api_client:
            sys.exit(1)
        # Always use AI_POD_master_Template.xlsx as the filename for setup action
        excel_file = "output/AI_POD_master_Template.xlsx"
        print_info(f"Using standard filename: {excel_file}")
        setup_excel_file(api_client, excel_file)
    elif args.action == 'get-info':
        api_client = get_api_client()
        if not api_client:
            sys.exit(1)
        get_intersight_info(api_client, args.file)
    else:
        api_client = get_api_client()
        if not api_client:
            sys.exit(1)
        
        # Automatically retrieve organization and server information first
        print('\n--- Automatically retrieving organization and server information from Intersight ---')
        get_intersight_info(api_client, args.file)
        print('--- Finished retrieving Intersight information ---\n')
        
        if args.action in ['push', 'all']:
            process_foundation_template(args.file)
        
        if args.action in ['template', 'all']:
            create_server_template_from_excel(api_client, args.file)
        
        if args.action in ['profiles', 'all']:
            create_server_profiles_from_excel(api_client, args.file)
