#!/usr/bin/env python3
"""
Script to update organization dropdowns in all sheets based on Intersight data
"""
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv
import os
import intersight
from intersight.api_client import ApiClient
from intersight.configuration import Configuration
from intersight.api import organization_api

def get_api_client():
    """
    Create an Intersight API client using the API key file
    """
    try:
        # Get API key details from environment variables
        api_key_id = os.getenv('INTERSIGHT_API_KEY_ID')
        api_key_file = os.getenv('INTERSIGHT_PRIVATE_KEY_FILE', './SecretKey.txt')
        
        if not api_key_id or not os.path.exists(api_key_file):
            print("Error: API key configuration not found")
            return None
            
        # Create configuration
        config = Configuration(
            host = os.getenv('INTERSIGHT_BASE_URL', 'https://intersight.com'),
            signing_info = intersight.signing.HttpSigningConfiguration(
                key_id = api_key_id,
                private_key_path = api_key_file,
                signing_scheme = intersight.signing.SCHEME_HS2019,
                signing_algorithm = intersight.signing.ALGORITHM_ECDSA_MODE_FIPS_186_3,
                hash_algorithm = intersight.signing.HASH_SHA256,
                signed_headers = [
                    intersight.signing.HEADER_REQUEST_TARGET,
                    intersight.signing.HEADER_HOST,
                    intersight.signing.HEADER_DATE,
                    intersight.signing.HEADER_DIGEST,
                ]
            )
        )
        
        # Create API client
        api_client = ApiClient(configuration=config)
        return api_client
        
    except Exception as e:
        print(f"Error creating API client: {str(e)}")
        return None

def update_organization_dropdowns(excel_file):
    """Update organization dropdowns in all sheets"""
    # Load environment variables
    load_dotenv()
    
    # Get API client
    api_client = get_api_client()
    if not api_client:
        sys.exit(1)
    
    # Get organizations from Intersight
    print("\nGetting organizations from Intersight...")
    org_api = organization_api.OrganizationApi(api_client)
    orgs = org_api.get_organization_organization_list()
    org_names = [org.name for org in orgs.results]
    print(f"Found {len(org_names)} organizations: {org_names}")
    
    # Load workbook
    try:
        workbook = load_workbook(excel_file)
        
        # Update Policies sheet organization dropdown
        if 'Policies' in workbook.sheetnames:
            policies_sheet = workbook['Policies']
            
            # Check if organization validation exists
            org_validation_exists = False
            for dv in list(policies_sheet.data_validations.dataValidation):
                for cell_range in dv.sqref:
                    if cell_range.startswith('D'):
                        # Found organization dropdown, update formula
                        dv.formula1 = f'"{",".join(org_names)}"'
                        org_validation_exists = True
                        break
            
            # If no organization validation exists, create one
            if not org_validation_exists:
                policies_org_validation = DataValidation(
                    type='list',
                    formula1=f'"{",".join(org_names)}"',
                    allow_blank=True
                )
                policies_org_validation.add('D2:D1000')  # Apply to Organization column
                policies_sheet.add_data_validation(policies_org_validation)
            
            print("Updated organization dropdown in Policies sheet")
        
        # Update Template sheet organization dropdown
        if 'Template' in workbook.sheetnames:
            template_sheet = workbook['Template']
            
            # Check if organization validation exists
            org_validation_exists = False
            for dv in list(template_sheet.data_validations.dataValidation):
                for cell_range in dv.sqref:
                    if cell_range.startswith('B'):
                        # Found organization dropdown, update formula
                        dv.formula1 = f'"{",".join(org_names)}"'
                        org_validation_exists = True
                        break
            
            # If no organization validation exists, create one
            if not org_validation_exists:
                template_org_validation = DataValidation(
                    type='list',
                    formula1=f'"{",".join(org_names)}"',
                    allow_blank=True
                )
                template_org_validation.add('B2:B1000')  # Apply to Organization column
                template_sheet.add_data_validation(template_org_validation)
            
            print("Updated organization dropdown in Template sheet")
        
        # Update Profiles sheet organization dropdown
        if 'Profiles' in workbook.sheetnames:
            profiles_sheet = workbook['Profiles']
            
            # Check if organization validation exists
            org_validation_exists = False
            for dv in list(profiles_sheet.data_validations.dataValidation):
                for cell_range in dv.sqref:
                    if cell_range.startswith('C'):
                        # Found organization dropdown, update formula
                        dv.formula1 = f'"{",".join(org_names)}"'
                        org_validation_exists = True
                        break
            
            # If no organization validation exists, create one
            if not org_validation_exists:
                profiles_org_validation = DataValidation(
                    type='list',
                    formula1=f'"{",".join(org_names)}"',
                    allow_blank=True
                )
                profiles_org_validation.add('C2:C1000')  # Apply to Organization column
                profiles_sheet.add_data_validation(profiles_org_validation)
            
            print("Updated organization dropdown in Profiles sheet")
        
        # Save the workbook
        workbook.save(excel_file)
        print(f"\nSuccessfully updated organization dropdowns in {excel_file}")
        
    except Exception as e:
        print(f"Error updating organization dropdowns: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python update_org_dropdowns_fixed.py <excel_file>")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    update_organization_dropdowns(excel_file)
