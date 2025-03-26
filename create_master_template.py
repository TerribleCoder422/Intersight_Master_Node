#!/usr/bin/env python3
"""
Create a static master template for Intersight Foundation based on the Intersight_Foundation.xlsx file.
"""

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os
import shutil

def create_master_template(source_file, output_file):
    """
    Create a master template with all the necessary sheets and dropdowns
    properly configured in the exact columns requested.
    """
    # First check if the source file exists
    if not os.path.exists(source_file):
        print(f"Source file {source_file} not found. Creating from scratch.")
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    else:
        print(f"Loading source file: {source_file}")
        # Copy the file first to avoid modifying the original
        shutil.copy(source_file, output_file)
        wb = load_workbook(output_file)
    
    # Current sheets
    current_sheets = wb.sheetnames
    print(f"Current sheets: {current_sheets}")
    
    # Remove any sheets with '1' suffix (duplicates)
    duplicate_sheets = [sheet for sheet in current_sheets if sheet.endswith('1')]
    for sheet_name in duplicate_sheets:
        print(f"Removing duplicate sheet: {sheet_name}")
        wb.remove(wb[sheet_name])
    
    # Define essential sheets in the desired order
    essential_sheets = ['Pools', 'Policies', 'Template', 'Profiles', 'Templates', 'Organizations', 'Servers']
    
    # Create missing sheets
    for sheet_name in essential_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Creating missing sheet: {sheet_name}")
            wb.create_sheet(sheet_name)
    
    # Create or update the Pools sheet
    if 'Pools' in wb.sheetnames:
        pools_sheet = wb['Pools']
        # Set up headers if the sheet is empty
        if pools_sheet.max_row <= 1:
            headers = ["Pool Type*", "Pool Name*", "Description", "Start Address*", "Size*"]
            for col, header in enumerate(headers, 1):
                cell = pools_sheet.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Add sample pool data
            sample_pools = [
                ("MAC Pool", "Ai_POD-MAC-A", "MAC Pool for AI POD Fabric A", "00:25:B5:A0:00:00", "256"),
                ("MAC Pool", "Ai_POD-MAC-B", "MAC Pool for AI POD Fabric B", "00:25:B5:B0:00:00", "256"),
                ("UUID Pool", "Ai_POD-UUID-Pool", "UUID Pool for AI POD Servers", "0000-000000000001", "100")
            ]
            for row_idx, row_data in enumerate(sample_pools, 2):
                for col_idx, value in enumerate(row_data, 1):
                    pools_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Create or update the Policies sheet
    if 'Policies' in wb.sheetnames:
        policies_sheet = wb['Policies']
        # Set up headers if the sheet is empty
        if policies_sheet.max_row <= 1:
            headers = ["Policy Type*", "Policy Name*", "Description", "Organization*"]
            for col, header in enumerate(headers, 1):
                cell = policies_sheet.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Add sample policy data
            sample_policies = [
                ('vNIC', 'Ai_POD-vNIC-A', 'vNIC Policy for AI POD Fabric A', 'default'),
                ('vNIC', 'Ai_POD-vNIC-B', 'vNIC Policy for AI POD Fabric B', 'default'),
                ('BIOS', 'Ai_POD-BIOS', 'BIOS Policy for AI POD', 'default'),
                ('BOOT', 'Ai_POD-BOOT', 'Boot Policy for AI POD', 'default'),
                ('QoS', 'Ai_POD-QoS', 'QoS Policy for AI POD', 'default'),
                ('Storage', 'Ai_POD-Storage', 'Storage Policy for AI POD', 'default')
            ]
            for row_idx, row_data in enumerate(sample_policies, 2):
                for col_idx, value in enumerate(row_data, 1):
                    policies_sheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Add organization dropdown to column D in Policies sheet
        org_validation_policies = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
        policies_sheet.add_data_validation(org_validation_policies)
        org_validation_policies.add('D2:D1000')  # Column D
        print("Added organization dropdown to column D in Policies sheet")
    
    # Create or update the Template sheet
    if 'Template' in wb.sheetnames:
        template_sheet = wb['Template']
        # Set up headers if the sheet is empty
        if template_sheet.max_row <= 1:
            template_headers = [
                "Template Name*", 
                "Organization*", 
                "Description",
                "Target Platform*",
                "BIOS Policy*",
                "Boot Policy*",
                "LAN Connectivity Policy*",
                "Storage Policy*"
            ]
            for col, header in enumerate(template_headers, 1):
                cell = template_sheet.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Add sample template data
            template_example = [
                "Ai_POD_Template",
                "default",
                "Server template for AI POD workloads",
                "FIAttached",
                "Ai_POD-BIOS",
                "Ai_POD-BOOT",
                "Ai_POD-vNIC-A",
                "Ai_POD-Storage"
            ]
            for col, value in enumerate(template_example, 1):
                template_sheet.cell(row=2, column=col, value=value)
        
        # Add organization dropdown to column B in Template sheet
        org_validation_template = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
        template_sheet.add_data_validation(org_validation_template)
        org_validation_template.add('B2:B1000')  # Column B
        print("Added organization dropdown to column B in Template sheet")
        
        # Add Target Platform dropdown to column D in Template sheet
        platform_validation = DataValidation(type='list', formula1='"FIAttached,Standalone"', allow_blank=True)
        template_sheet.add_data_validation(platform_validation)
        platform_validation.add('D2:D1000')  # Column D
        print("Added Target Platform dropdown to column D in Template sheet")
    
    # Create or update the Profiles sheet
    if 'Profiles' in wb.sheetnames:
        profiles_sheet = wb['Profiles']
        # Set up headers if the sheet is empty
        if profiles_sheet.max_row <= 1:
            profile_headers = ["Profile Name*", "Description", "Organization*", "Template Name*", "Server*", "Description", "Deploy*"]
            for col, header in enumerate(profile_headers, 1):
                cell = profiles_sheet.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)
            
            # Add sample profile data
            profiles_sheet.append(['AI-Server-01', 'AI POD Host Profile', 'default', 'Ai_Pod_Template', '', 'AI POD Server Profile', 'No'])
        
        # Add organization dropdown to column C in Profiles sheet
        org_validation_profiles = DataValidation(type='list', formula1='"default,DevOps,Production,Test,UAT"', allow_blank=True)
        profiles_sheet.add_data_validation(org_validation_profiles)
        org_validation_profiles.add('C2:C1000')  # Column C
        print("Added organization dropdown to column C in Profiles sheet")
        
        # Add server dropdown to column E in Profiles sheet
        server_validation = DataValidation(type='list', formula1='"Server-1 (FCH1234V5Z7),Server-2 (FCH5678A9BC),Server-3 (FCH9012D3EF)"', allow_blank=True)
        profiles_sheet.add_data_validation(server_validation)
        server_validation.add('E2:E1000')  # Column E
        print("Added server dropdown to column E in Profiles sheet")
        
        # Add deploy dropdown validation
        deploy_validation = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
        profiles_sheet.add_data_validation(deploy_validation)
        deploy_validation.add('G2:G1000')
        print("Added deploy dropdown to column G in Profiles sheet")
    
    # Set column widths for all sheets
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"Master template saved as: {output_file}")
    return True

if __name__ == "__main__":
    source_file = "Intersight_Foundation.xlsx"
    output_file = "Intersight_Master_Node.xlsx"
    create_master_template(source_file, output_file)
