import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Initialize policy data
def initialize_policy_data():
    policies_data = {
        'Name': [],
        'Description': [],
        'Organization': [],
        'Policy Type': [],
        # Add specific settings for each policy type
        'BIOS Settings': [],  # For BIOS policies
        'Boot Devices': [],   # For Boot policies
        'VLAN Settings': [],  # For vNIC policies
        'VSAN Settings': [],  # For vHBA policies
        'Media Type': [],     # For vMedia policies
        'RAID Level': [],     # For Storage policies
        'QoS Priority': []    # For QoS policies
    }
    return policies_data

# Add predefined policies with their specific settings
def add_policies(policies_data):
    policies = [
        ('QoS-Policy-Test', 'High priority QoS policy', 'default', 'QoS', '', '', '', '', '', '', 'Gold'),
        ('Boot-Policy-Test', 'Boot policy for servers', 'default', 'Boot', '', 'M2,LocalDisk,vMedia', '', '', '', '', ''),
        ('vNIC-Policy-Test', 'vNIC settings', 'default', 'vNIC', '', '', 'VLAN-1,VLAN-2', '', '', '', ''),
        ('vHBA-Policy-Test', 'vHBA settings', 'default', 'vHBA', '', '', '', 'VSAN-100', '', '', ''),
        ('BIOS-Policy-Test', 'BIOS configuration', 'default', 'BIOS', 'Performance', '', '', '', '', '', ''),
        ('vMedia-Policy-Test', 'vMedia boot settings', 'default', 'vMedia', '', '', '', '', 'ISO', '', ''),
        ('Storage-Policy-Test', 'Storage configuration', 'default', 'Storage', '', '', '', '', '', 'RAID-1', '')
    ]
    
    for policy in policies:
        for i, key in enumerate(policies_data.keys()):
            policies_data[key].append(policy[i])

# Create Excel template with policy-specific sheets
def create_excel_template(output_file, policies_data):
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    policies_df = pd.DataFrame(policies_data)
    policies_df.to_excel(writer, sheet_name='Policies', index=False, startrow=2)

    workbook = writer.book
    policies_sheet = writer.sheets['Policies']

    # Add title and instructions
    title_cell = policies_sheet.cell(row=1, column=1, value="Intersight Policy Configuration Template")
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='000000')
    policies_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(policies_data.keys()))
    
    # Style constants
    COLORS = {
        'header_bg': '1F4E78',  # Dark blue
        'header_font': 'FFFFFF',  # White
        'subheader_bg': '4472C4',  # Medium blue
        'alt_row_bg': 'F2F2F2',  # Light gray
        'border': '8EA9DB'  # Light blue
    }

    # Create styles
    header_fill = PatternFill(start_color=COLORS['header_bg'], end_color=COLORS['header_bg'], fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color=COLORS['header_font'])
    alt_row_fill = PatternFill(start_color=COLORS['alt_row_bg'], end_color=COLORS['alt_row_bg'], fill_type='solid')
    border = Border(
        left=Side(style='thin', color=COLORS['border']),
        right=Side(style='thin', color=COLORS['border']),
        top=Side(style='thin', color=COLORS['border']),
        bottom=Side(style='thin', color=COLORS['border'])
    )

    # Style and adjust all columns
    for col in range(1, len(policies_data.keys()) + 1):
        column_letter = get_column_letter(col)
        
        # Style header cell
        header_cell = policies_sheet.cell(row=3, column=col)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.border = border
        
        # Set wider default column width and enable text wrapping
        policies_sheet.column_dimensions[column_letter].width = 30  # Set default width to 30 characters
        
        # Apply styling to all cells in the column
        for row in range(1, policies_sheet.max_row + 1):
            cell = policies_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, 
                                    vertical='center',
                                    horizontal='left',
                                    shrink_to_fit=False)  # Prevent text shrinking
            
            if row > 3:  # Skip header rows
                cell.border = border
                if row % 2 == 0:
                    cell.fill = alt_row_fill

    # Freeze panes
    policies_sheet.freeze_panes = 'A4'

    # Add data validation for Policy Type with wider dropdown
    policy_types = "QoS,Boot,vNIC,vHBA,BIOS,vMedia,Storage"
    policy_type_dv = DataValidation(
        type="list",
        formula1=f'"{policy_types}"',
        allow_blank=True,
        showDropDown=True,
        showInputMessage=True,
        promptTitle='Policy Type',
        prompt='Select the type of policy to configure'
    )
    policies_sheet.add_data_validation(policy_type_dv)
    policy_col = get_column_letter(policies_df.columns.get_loc('Policy Type') + 1)
    for row in range(4, len(policies_data['Policy Type']) + 4):
        policy_type_dv.add(f'{policy_col}{row}')

    # Add data validation for specific settings
    qos_priority_dv = DataValidation(
        type="list",
        formula1='"Platinum,Gold,Silver,Bronze"',
        allow_blank=True,
        showDropDown=True,
        showInputMessage=True,
        promptTitle='QoS Priority',
        prompt='Select the QoS priority level'
    )
    policies_sheet.add_data_validation(qos_priority_dv)
    
    raid_level_dv = DataValidation(
        type="list",
        formula1='"RAID-0,RAID-1,RAID-5,RAID-6,RAID-10"',
        allow_blank=True,
        showDropDown=True,
        showInputMessage=True,
        promptTitle='RAID Level',
        prompt='Select the RAID configuration level'
    )
    policies_sheet.add_data_validation(raid_level_dv)

    # Create reference sheet
    ref_sheet = workbook.create_sheet('Reference')
    ref_data = {
        'BIOS Presets': ['Performance', 'LowLatency', 'Custom'],
        'Boot Device Types': ['LocalDisk', 'M2', 'vMedia', 'PXE', 'iSCSI', 'SAN'],
        'Media Types': ['ISO', 'IMG', 'HTTP', 'CIFS'],
        'RAID Levels': ['RAID-0', 'RAID-1', 'RAID-5', 'RAID-6', 'RAID-10'],
        'QoS Priorities': ['Platinum', 'Gold', 'Silver', 'Bronze']
    }

    # Style the reference sheet
    col = 1
    for category, values in ref_data.items():
        column_letter = get_column_letter(col)
        ref_sheet.column_dimensions[column_letter].width = 30  # Set wider default width
        
        # Add category header
        header_cell = ref_sheet.cell(row=1, column=col, value=category)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.border = border
        header_cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Add values
        for i, value in enumerate(values, start=2):
            cell = ref_sheet.cell(row=i, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if i % 2 == 0:
                cell.fill = alt_row_fill
        
        col += 1

    writer.close()
    print(f"Excel template has been created at: {output_file}")

if __name__ == "__main__":
    output_file = 'output/Intersight_Policies.xlsx'
    os.makedirs('output', exist_ok=True)
    policies_data = initialize_policy_data()
    add_policies(policies_data)
    create_excel_template(output_file, policies_data)
