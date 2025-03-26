#!/usr/bin/env python3
"""
Simpler script to fix the Policies sheet mixup in the Intersight Excel template
"""

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill

def fix_excel_template(file_path):
    """Fix Intersight Excel template by transferring data to the correct sheets."""
    print(f"Opening workbook: {file_path}")
    wb = load_workbook(file_path)
    
    # Check for both Policies and Policies1 sheets
    print(f"Current sheets: {wb.sheetnames}")
    
    # Handle the Policies/Policies1 mix-up
    if 'Policies' in wb.sheetnames and 'Policies1' in wb.sheetnames:
        # Copy content from Policies1 to Policies since Policies1 has the data
        policies1 = wb['Policies1']
        policies = wb['Policies']
        
        # Clear the Policies sheet first
        for row in list(policies.rows):
            for cell in row:
                cell.value = None
        
        # Copy just the values from Policies1 to Policies
        for row_idx, row in enumerate(policies1.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                policies.cell(row=row_idx, column=col_idx).value = cell.value
                
                # Add basic styling for headers
                if row_idx == 1:
                    policies.cell(row=row_idx, column=col_idx).font = Font(color='FFFFFF', bold=True)
                    policies.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
        
        # Now remove the Policies1 sheet
        wb.remove(policies1)
    
    # Make sure we have all duplicate sheets too
    duplicate_sheets = []
    for sheet in wb.sheetnames:
        if sheet.endswith('1'):
            duplicate_sheets.append(sheet)
    
    # Remove any remaining duplicate sheets
    for dup_sheet in duplicate_sheets:
        if dup_sheet in wb.sheetnames:
            print(f"Removing duplicate sheet: {dup_sheet}")
            wb.remove(wb[dup_sheet])
    
    # Create sample lists for dropdowns
    org_list = ["default", "DevOps", "Production", "Test", "UAT"]
    server_list = ["Server-1 (FCH1234V5Z7)", "Server-2 (FCH5678A9BC)", "Server-3 (FCH9012D3EF)"]
    
    # Profiles sheet: Organization in column C, Server dropdown in column E
    if 'Profiles' in wb.sheetnames:
        profiles_sheet = wb['Profiles']
        
        # Add organization dropdown to column C
        org_validation_profiles = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        profiles_sheet.add_data_validation(org_validation_profiles)
        org_validation_profiles.add('C2:C1000')  # Column C
        
        # Add server dropdown to column E
        server_validation = DataValidation(type='list', formula1=f'"{",".join(server_list)}"', allow_blank=True)
        profiles_sheet.add_data_validation(server_validation)
        server_validation.add('E2:E1000')  # Column E
    
    # Policies sheet: Organization in column D
    if 'Policies' in wb.sheetnames:
        policies_sheet = wb['Policies']
        
        # Add organization dropdown to column D
        org_validation_policies = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        policies_sheet.add_data_validation(org_validation_policies)
        org_validation_policies.add('D2:D1000')  # Column D
    
    # Template sheet: Organization in column B
    if 'Template' in wb.sheetnames:
        template_sheet = wb['Template']
        
        # Add organization dropdown to column B
        org_validation_template = DataValidation(type='list', formula1=f'"{",".join(org_list)}"', allow_blank=True)
        template_sheet.add_data_validation(org_validation_template)
        org_validation_template.add('B2:B1000')  # Column B
    
    # Save the fixed workbook
    fixed_file_path = file_path.replace('.xlsx', '_fixed.xlsx')
    wb.save(fixed_file_path)
    print(f"Fixed workbook saved to: {fixed_file_path}")
    return fixed_file_path

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        file_path = 'intersight_simplified_template.xlsx'
    
    fix_excel_template(file_path)
